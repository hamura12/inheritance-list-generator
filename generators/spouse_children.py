"""配偶者・子（１人～４人まで対応）である場合"""
from openpyxl import Workbook
from .common import (F11, F14, YEL, al, set_col_widths, set_row_heights,
                     merge, fill_yellow, set_border, set_cell, draw_creator_box)


def generate(num_children: int,
             appl_start: str = None, appl_end: str = None) -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)
    _build_sheet(wb, num_children, appl_start, appl_end)
    return wb


def _child_start(n: int) -> int:
    """子n（n>=4）のRIGHT側開始行"""
    return 31 + (n - 4) * 6


def _add_child_block_right(ws, start: int):
    """子4以降のRIGHT側ブロック（6行ひとまとまり）"""
    merge(ws, f'P{start}', f'Q{start}', '住所', alignment=al(1))
    fill_yellow(ws, start, 18, start, 31)
    merge(ws, f'R{start}', f'AE{start}')
    merge(ws, f'P{start+1}', f'Q{start+1}', '出生', alignment=al(1))
    fill_yellow(ws, start+1, 18, start+1, 25)
    merge(ws, f'R{start+1}', f'Y{start+1}')
    fill_yellow(ws, start+2, 16, start+2, 18)
    merge(ws, f'P{start+2}', f'R{start+2}', '（　　）', alignment=al(1))
    fill_yellow(ws, start+3, 16, start+4, 24)
    merge(ws, f'P{start+3}', f'X{start+4}')


def _add_m_borders_for_child(ws, start: int):
    """M列の区切り罫線（子4以降のセクション）"""
    set_border(ws, start-2, 13, top=1, left=1)
    set_border(ws, start-2, 14, top=1)
    set_border(ws, start-2, 15, top=1)
    for r in range(start-1, start+3):
        set_border(ws, r, 13, left=1)
    set_border(ws, start+3, 13, bottom=1, left=1)
    set_border(ws, start+3, 14, bottom=1)
    set_border(ws, start+3, 15, bottom=1)


def _build_sheet(wb, num_children, appl_start=None, appl_end=None):
    # デフォルト申出人位置（配偶者欄）
    if appl_start is None:
        appl_start = 'Y16' if num_children == 1 else 'Y13'
        appl_end   = 'AB17' if num_children == 1 else 'AB14'
    if num_children <= 4:
        titles = {1: '配偶者・子1人の場合', 2: '配偶者・子2人の場合',
                  3: '配偶者・子3人の場合', 4: '配偶者・子4人の場合'}
        ws = wb.create_sheet(titles[num_children])
        set_col_widths(ws)
        if num_children == 1:
            _sheet1(ws)
        elif num_children == 2:
            _sheet2(ws)
        elif num_children == 3:
            _sheet3(ws)
        elif num_children == 4:
            _sheet4(ws)
        # 申出人ラベルを選択した相続人の氏名欄右に配置
        merge(ws, appl_start, appl_end, '(申出人)', alignment=al(2))
    else:
        _sheet_extended(wb, num_children, appl_start, appl_end)


def _sheet1(ws):
    set_row_heights(ws, {4:585,5:90,6:300,7:300,8:300,9:300,10:300,11:150,12:150,
        13:300,14:300,15:300,16:150,17:150,18:150,19:150,20:300,21:300,
        22:150,23:150,24:300,26:150,27:300,28:300,29:300,30:300,
        32:300,33:300,34:300,35:300,36:300,37:300,38:300,39:300})

    merge(ws,'I4','M4','被相続人',F14,alignment=al(3))
    fill_yellow(ws,4,14,4,20); merge(ws,'N4','T4')
    merge(ws,'U4','AA4','法定相続情報',F14,alignment=al(2))

    merge(ws,'A6','E6','最後の住所',alignment=al(1))
    fill_yellow(ws,7,1,7,14); merge(ws,'A7','N7')
    merge(ws,'A8','E8','最後の本籍',alignment=al(1))
    fill_yellow(ws,9,1,9,14); merge(ws,'A9','N9')
    merge(ws,'P9','Q9')
    fill_yellow(ws,9,18,9,27); merge(ws,'R9','AA9')
    merge(ws,'A10','B10','出生  ',alignment=al(1))
    fill_yellow(ws,10,3,10,11); merge(ws,'C10','K10')
    merge(ws,'A11','B12','死亡  ',alignment=al(1))
    fill_yellow(ws,11,3,12,11); merge(ws,'C11','K12')

    merge(ws,'A13','E13','（被相続人）',alignment=al(2))
    merge(ws,'P13','Q13','住所',alignment=al(1))
    fill_yellow(ws,13,18,13,31); merge(ws,'R13','AE13')
    fill_yellow(ws,14,1,14,9); merge(ws,'A14','I14')
    merge(ws,'P14','Q14','出生  ',alignment=al(1))
    fill_yellow(ws,14,18,14,27); merge(ws,'R14','AA14')

    fill_yellow(ws,15,16,15,18); merge(ws,'P15','R15','（　　）',alignment=al(1))
    fill_yellow(ws,16,16,17,24); merge(ws,'P16','X17')
    fill_yellow(ws,18,16,19,24); merge(ws,'P18','X19')

    merge(ws,'A20','B20','住所　',alignment=al(1))
    fill_yellow(ws,20,3,20,15); merge(ws,'C20','O20')
    merge(ws,'P20','Q20')
    fill_yellow(ws,20,18,20,27); merge(ws,'R20','AA20')
    merge(ws,'A21','B21','出生  ',alignment=al(0))
    fill_yellow(ws,21,3,21,11); merge(ws,'C21','K21')
    fill_yellow(ws,22,1,23,5); merge(ws,'A22','E23','（　　）',alignment=al(1))
    fill_yellow(ws,22,16,23,24); merge(ws,'P22','X23')
    fill_yellow(ws,24,1,24,9); merge(ws,'A24','I24')
    merge(ws,'P24','T24','以下余白',alignment=al(2))

    set_cell(ws,27,11,'作成日：',alignment=al(0))
    fill_yellow(ws,27,14,27,24); merge(ws,'N27','X27')
    set_cell(ws,28,11,'作成者：',alignment=al(0))
    merge(ws,'N28','O28','住所',alignment=al(1))
    fill_yellow(ws,28,16,28,28); merge(ws,'P28','AB28')
    set_cell(ws,29,14,'氏名',alignment=al(2))
    fill_yellow(ws,29,16,29,22); merge(ws,'P29','V29')

    set_border(ws,15,3,left=6); set_border(ws,16,3,left=6)
    set_border(ws,16,13,bottom=1); set_border(ws,16,14,bottom=1); set_border(ws,16,15,bottom=1)
    set_border(ws,17,3,top=1,left=6)
    for c in range(4,16): set_border(ws,17,c,top=1)
    set_border(ws,18,3,left=6); set_border(ws,19,3,left=6)
    draw_creator_box(ws, 26)


def _sheet2(ws):
    set_row_heights(ws, {4:585,5:90,6:90,7:300,8:300,9:300,10:300,11:150,12:150,13:150,14:150,
        15:300,16:300,17:300,18:150,19:150,20:150,21:150,22:300,23:300,
        24:150,25:150,26:300,27:300,29:150,30:300,31:300,32:300,33:300,
        35:300,36:300,37:300,38:300,39:300,40:300,41:300,42:300})

    merge(ws,'I4','M4','被相続人',F14,alignment=al(3))
    fill_yellow(ws,4,14,4,20); merge(ws,'N4','T4')
    merge(ws,'U4','AA4','法定相続情報',F14,alignment=al(2))

    merge(ws,'A7','K7','最後の住所',alignment=al(1))
    fill_yellow(ws,8,1,8,13); merge(ws,'A8','M8')
    merge(ws,'A9','K9','最後の本籍　',alignment=al(1))
    merge(ws,'P9','Q9','住所',alignment=al(1))
    fill_yellow(ws,9,18,9,31); merge(ws,'R9','AE9')
    fill_yellow(ws,10,1,10,13); merge(ws,'A10','M10')
    merge(ws,'P10','Q10','出生  ',alignment=al(1))
    fill_yellow(ws,10,18,10,27); merge(ws,'R10','AA10')

    merge(ws,'A11','B12','出生  ',alignment=al(1))
    fill_yellow(ws,11,3,12,11); merge(ws,'C11','K12')
    fill_yellow(ws,11,16,12,18); merge(ws,'P11','R12','（　　）',alignment=al(1))

    merge(ws,'A13','B14','死亡  ',alignment=al(1))
    fill_yellow(ws,13,3,14,11); merge(ws,'C13','K14')
    fill_yellow(ws,13,16,14,24); merge(ws,'P13','X14')


    merge(ws,'A15','E15','（被相続人）',alignment=al(2))
    fill_yellow(ws,16,1,16,9); merge(ws,'A16','I16')

    merge(ws,'P20','Q21','住所',alignment=al(1))
    fill_yellow(ws,20,18,21,31); merge(ws,'R20','AE21')
    merge(ws,'A22','B22','住所　',alignment=al(1))
    fill_yellow(ws,22,3,22,13); merge(ws,'C22','M22')
    merge(ws,'P22','Q22','出生  ',alignment=al(1))
    fill_yellow(ws,22,18,22,27); merge(ws,'R22','AA22')
    merge(ws,'A23','B23','出生  ',alignment=al(0))
    fill_yellow(ws,23,3,23,11); merge(ws,'C23','K23')
    fill_yellow(ws,23,16,23,18); merge(ws,'P23','R23','（　　）',alignment=al(1))

    fill_yellow(ws,24,1,25,5); merge(ws,'A24','E25','（　　）',alignment=al(1))
    fill_yellow(ws,24,16,25,24); merge(ws,'P24','X25')
    fill_yellow(ws,26,1,26,9); merge(ws,'A26','I26')
    merge(ws,'P27','T27','以下余白',alignment=al(2))

    set_cell(ws,30,11,'作成日：',alignment=al(0))
    fill_yellow(ws,30,14,30,24); merge(ws,'N30','X30')
    set_cell(ws,31,11,'作成者：',alignment=al(0))
    merge(ws,'N31','O31','住所',alignment=al(1))
    fill_yellow(ws,31,16,31,28); merge(ws,'P31','AB31')
    set_cell(ws,32,14,'氏名',alignment=al(2))
    fill_yellow(ws,32,16,32,22); merge(ws,'P32','V32')

    set_border(ws,14,13,right=1); set_border(ws,14,14,top=1,left=1); set_border(ws,14,15,top=1)
    for r in range(15,25):
        set_border(ws,r,13,right=1); set_border(ws,r,14,left=1)
    set_border(ws,17,3,left=6); set_border(ws,18,3,left=6)
    set_border(ws,19,3,top=1,left=6)
    for c in range(4,13): set_border(ws,19,c,top=1)
    set_border(ws,19,13,top=1,right=1)
    set_border(ws,20,3,left=6); set_border(ws,21,3,left=6)
    set_border(ws,24,13,right=1); set_border(ws,24,14,top=1,left=1); set_border(ws,24,15,top=1)
    draw_creator_box(ws, 29)


def _sheet3(ws):
    set_row_heights(ws, {4:585,5:90,6:300,7:300,8:300,9:300,10:300,11:150,12:150,13:150,14:150,
        15:300,16:300,17:150,18:150,19:150,20:150,21:150,22:150,23:300,
        24:150,25:150,26:300,27:300,28:150,29:150,30:300,31:300,33:150,
        34:300,35:300,36:300,37:300,39:300,40:300,41:300,42:300,43:300,
        44:300,45:300,46:300})

    merge(ws,'I4','M4','被相続人',F14,alignment=al(3))
    fill_yellow(ws,4,14,4,20); merge(ws,'N4','T4')
    merge(ws,'U4','AA4','法定相続情報',F14,alignment=al(2))

    merge(ws,'A6','E6','最後の住所',alignment=al(1))
    fill_yellow(ws,7,1,7,12); merge(ws,'A7','L7')
    merge(ws,'A8','E8','最後の本籍',alignment=al(1))
    fill_yellow(ws,9,1,9,12); merge(ws,'A9','L9')
    merge(ws,'P9','Q9','住所',alignment=al(1))
    fill_yellow(ws,9,18,9,31); merge(ws,'R9','AE9')
    merge(ws,'A10','B10','出生  ',alignment=al(1))
    fill_yellow(ws,10,3,10,11); merge(ws,'C10','K10')
    merge(ws,'P10','Q10','出生  ',alignment=al(1))
    fill_yellow(ws,10,18,10,25); merge(ws,'R10','Y10')
    merge(ws,'A11','B12','死亡  ',alignment=al(1))
    fill_yellow(ws,11,3,12,11); merge(ws,'C11','K12')
    fill_yellow(ws,11,16,12,18); merge(ws,'P11','R12','（　　）',alignment=al(1))
    fill_yellow(ws,11,25,12,28); merge(ws,'Y11','AB12')

    fill_yellow(ws,13,1,14,5); merge(ws,'A13','E14','（被相続人）',alignment=al(1))
    fill_yellow(ws,13,16,14,24); merge(ws,'P13','X14')

    fill_yellow(ws,15,1,15,9); merge(ws,'A15','I15')

    merge(ws,'P16','Q16','住所',alignment=al(1))
    fill_yellow(ws,16,18,16,31); merge(ws,'R16','AE16')
    merge(ws,'P17','Q18','出生',alignment=al(1))
    fill_yellow(ws,17,18,18,25); merge(ws,'R17','Y18')
    fill_yellow(ws,19,16,20,18); merge(ws,'P19','R20','（　　）',alignment=al(1))
    fill_yellow(ws,21,1,22,2); merge(ws,'A21','B22','住所　',alignment=al(1))
    fill_yellow(ws,21,3,22,12); merge(ws,'C21','L22')
    fill_yellow(ws,21,16,22,24); merge(ws,'P21','X22')
    merge(ws,'A23','B23','出生  ',alignment=al(0))
    fill_yellow(ws,23,3,23,11); merge(ws,'C23','K23')

    fill_yellow(ws,24,1,25,5); merge(ws,'A24','E25','（　　）',alignment=al(1))
    merge(ws,'P24','Q25','住所',alignment=al(1))
    fill_yellow(ws,24,18,25,31); merge(ws,'R24','AE25')
    fill_yellow(ws,26,1,26,9); merge(ws,'A26','I26')
    merge(ws,'P26','Q26','出生',alignment=al(1))
    fill_yellow(ws,26,18,26,25); merge(ws,'R26','Y26')
    fill_yellow(ws,27,16,27,18); merge(ws,'P27','R27','（　　）',alignment=al(1))
    fill_yellow(ws,28,16,29,24); merge(ws,'P28','X29')
    merge(ws,'P31','T31','以下余白',alignment=al(2))

    set_cell(ws,34,11,'作成日：',alignment=al(0))
    fill_yellow(ws,34,14,34,24); merge(ws,'N34','X34')
    set_cell(ws,35,11,'作成者：',alignment=al(0))
    merge(ws,'N35','O35','住所',alignment=al(1))
    fill_yellow(ws,35,16,35,28); merge(ws,'P35','AB35')
    set_cell(ws,36,14,'氏名',alignment=al(2))
    fill_yellow(ws,36,16,36,22); merge(ws,'P36','V36')

    set_border(ws,13,13,bottom=1); set_border(ws,13,14,bottom=1); set_border(ws,13,15,bottom=1)
    set_border(ws,14,13,top=1,left=1); set_border(ws,14,14,top=1); set_border(ws,14,15,top=1)
    set_border(ws,15,13,left=1)
    set_border(ws,16,3,left=6); set_border(ws,16,13,left=1)
    set_border(ws,17,3,left=6); set_border(ws,17,13,left=1)
    set_border(ws,18,3,top=1,left=6)
    for c in range(4,13): set_border(ws,18,c,top=1)
    set_border(ws,18,13,left=1)
    set_border(ws,19,3,left=6); set_border(ws,19,13,left=1)
    set_border(ws,20,3,left=6); set_border(ws,20,13,left=1)
    set_border(ws,21,12,right=1); set_border(ws,21,13,left=1)
    set_border(ws,22,12,right=1)
    set_border(ws,22,13,top=1,left=1); set_border(ws,22,14,top=1); set_border(ws,22,15,top=1)
    for r in range(23,28): set_border(ws,r,13,left=1)
    set_border(ws,28,13,bottom=1,left=1); set_border(ws,28,14,bottom=1); set_border(ws,28,15,bottom=1)
    draw_creator_box(ws, 33)


def _sheet4(ws):
    set_row_heights(ws, {4:585,5:90,6:300,7:300,8:300,9:300,10:300,11:150,12:150,13:150,14:150,
        15:300,16:300,17:150,18:150,19:150,20:150,21:150,22:150,23:300,
        24:150,25:150,26:300,27:300,28:150,29:150,30:300,31:300,32:300,33:300,
        34:150,35:150,36:300,37:300,38:300,39:150,40:300,41:300,42:300,43:300,
        45:300,46:300,47:300,48:300,49:300,50:300,51:300,52:300})

    merge(ws,'I4','M4','被相続人',F14,alignment=al(3))
    fill_yellow(ws,4,14,4,20); merge(ws,'N4','T4')
    merge(ws,'U4','AA4','法定相続情報',F14,alignment=al(2))

    merge(ws,'A6','E6','最後の住所',alignment=al(1))
    fill_yellow(ws,7,1,7,12); merge(ws,'A7','L7')
    merge(ws,'A8','E8','最後の本籍',alignment=al(1))
    fill_yellow(ws,9,1,9,12); merge(ws,'A9','L9')
    merge(ws,'P9','Q9','住所',alignment=al(1))
    fill_yellow(ws,9,18,9,31); merge(ws,'R9','AE9')
    merge(ws,'A10','B10','出生  ',alignment=al(1))
    fill_yellow(ws,10,3,10,11); merge(ws,'C10','K10')
    merge(ws,'P10','Q10','出生  ',alignment=al(1))
    fill_yellow(ws,10,18,10,25); merge(ws,'R10','Y10')
    merge(ws,'A11','B12','死亡  ',alignment=al(1))
    fill_yellow(ws,11,3,12,11); merge(ws,'C11','K12')
    fill_yellow(ws,11,16,12,18); merge(ws,'P11','R12','（　　）',alignment=al(1))
    fill_yellow(ws,11,25,12,28); merge(ws,'Y11','AB12')

    fill_yellow(ws,13,1,14,5); merge(ws,'A13','E14','（被相続人）',alignment=al(1))
    fill_yellow(ws,13,16,14,24); merge(ws,'P13','X14')

    fill_yellow(ws,15,1,15,9); merge(ws,'A15','I15')
    merge(ws,'P15','Q15')
    fill_yellow(ws,15,18,15,27); merge(ws,'R15','AA15')

    merge(ws,'P16','Q16','住所',alignment=al(1))
    fill_yellow(ws,16,18,16,31); merge(ws,'R16','AE16')
    merge(ws,'P17','Q18','出生',alignment=al(1))
    fill_yellow(ws,17,18,18,25); merge(ws,'R17','Y18')
    fill_yellow(ws,19,16,20,18); merge(ws,'P19','R20','（　　）',alignment=al(1))
    fill_yellow(ws,21,1,22,2); merge(ws,'A21','B22','住所　',alignment=al(1))
    fill_yellow(ws,21,3,22,12); merge(ws,'C21','L22')
    fill_yellow(ws,21,16,22,24); merge(ws,'P21','X22')
    merge(ws,'A23','B23','出生  ',alignment=al(0))
    fill_yellow(ws,23,3,23,11); merge(ws,'C23','K23')

    fill_yellow(ws,24,1,25,5); merge(ws,'A24','E25','（　　）',alignment=al(1))
    merge(ws,'P24','Q25','住所',alignment=al(1))
    fill_yellow(ws,24,18,25,31); merge(ws,'R24','AE25')
    fill_yellow(ws,26,1,26,9); merge(ws,'A26','I26')
    merge(ws,'P26','Q26','出生',alignment=al(1))
    fill_yellow(ws,26,18,26,25); merge(ws,'R26','Y26')
    fill_yellow(ws,27,16,27,18); merge(ws,'P27','R27','（　　）',alignment=al(1))
    fill_yellow(ws,28,16,29,24); merge(ws,'P28','X29')

    merge(ws,'P31','Q31','住所',alignment=al(1))
    fill_yellow(ws,31,18,31,31); merge(ws,'R31','AE31')
    merge(ws,'P32','Q32','出生',alignment=al(1))
    fill_yellow(ws,32,18,32,25); merge(ws,'R32','Y32')
    fill_yellow(ws,33,16,33,18); merge(ws,'P33','R33','（　　）',alignment=al(1))
    fill_yellow(ws,34,16,35,24); merge(ws,'P34','X35')
    merge(ws,'P37','T37','以下余白',alignment=al(2))

    set_cell(ws,40,11,'作成日：',alignment=al(0))
    fill_yellow(ws,40,14,40,24); merge(ws,'N40','X40')
    set_cell(ws,41,11,'作成者：',alignment=al(0))
    merge(ws,'N41','O41','住所',alignment=al(1))
    fill_yellow(ws,41,16,41,28); merge(ws,'P41','AB41')
    set_cell(ws,42,14,'氏名',alignment=al(2))
    fill_yellow(ws,42,16,42,22); merge(ws,'P42','V42')

    set_border(ws,13,13,bottom=1); set_border(ws,13,14,bottom=1); set_border(ws,13,15,bottom=1)
    set_border(ws,14,13,top=1,left=1); set_border(ws,14,14,top=1); set_border(ws,14,15,top=1)
    set_border(ws,15,13,left=1)
    set_border(ws,16,3,left=6); set_border(ws,16,13,left=1)
    set_border(ws,17,3,left=6); set_border(ws,17,13,left=1)
    set_border(ws,18,3,top=1,left=6)
    for c in range(4,13): set_border(ws,18,c,top=1)
    set_border(ws,18,13,left=1)
    set_border(ws,19,3,left=6); set_border(ws,19,13,left=1)
    set_border(ws,20,3,left=6); set_border(ws,20,13,left=1)
    set_border(ws,21,12,right=1); set_border(ws,21,13,left=1)
    set_border(ws,22,12,right=1)
    set_border(ws,22,13,top=1,left=1); set_border(ws,22,14,top=1); set_border(ws,22,15,top=1)
    for r in range(23,28): set_border(ws,r,13,left=1)
    set_border(ws,28,13,bottom=1,left=1); set_border(ws,28,14,bottom=1); set_border(ws,28,15,bottom=1)
    set_border(ws,29,13,top=1,left=1); set_border(ws,29,14,top=1); set_border(ws,29,15,top=1)
    for r in range(30,34): set_border(ws,r,13,left=1)
    set_border(ws,34,13,bottom=1,left=1); set_border(ws,34,14,bottom=1); set_border(ws,34,15,bottom=1)
    for c in range(16,25): set_border(ws,38,c,bottom=1)
    draw_creator_box(ws, 39)


def _sheet_extended(wb, num_children, appl_start='Y13', appl_end='AB14'):
    """5人以上の子に対応した拡張シート"""
    ws = wb.create_sheet(f'配偶者・子{num_children}人の場合')
    set_col_widths(ws)

    # 行高さ（子4分のベース + 子5以降の各ブロック + 以下余白 + 作成者セクション）
    heights = {4:585,5:90,6:300,7:300,8:300,9:300,10:300,
        11:150,12:150,13:150,14:150,15:300,16:300,
        17:150,18:150,19:150,20:150,21:150,22:150,23:300,
        24:150,25:150,26:300,27:300,28:150,29:150,30:300}
    for n in range(4, num_children + 1):
        s = _child_start(n)
        heights[s] = 300; heights[s+1] = 300; heights[s+2] = 300
        heights[s+3] = 150; heights[s+4] = 150; heights[s+5] = 300
    wr = _child_start(num_children) + 6  # 以下余白 行
    heights[wr] = 300; heights[wr+1] = 300; heights[wr+2] = 150
    heights[wr+3] = 300; heights[wr+4] = 300; heights[wr+5] = 300
    set_row_heights(ws, heights)

    # ヘッダー（行4）
    merge(ws,'I4','M4','被相続人',F14,alignment=al(3))
    fill_yellow(ws,4,14,4,20); merge(ws,'N4','T4')
    merge(ws,'U4','AA4','法定相続情報',F14,alignment=al(2))

    # 被相続人情報（_sheet4と同一）
    merge(ws,'A6','E6','最後の住所',alignment=al(1))
    fill_yellow(ws,7,1,7,12); merge(ws,'A7','L7')
    merge(ws,'A8','E8','最後の本籍',alignment=al(1))
    fill_yellow(ws,9,1,9,12); merge(ws,'A9','L9')
    merge(ws,'P9','Q9','住所',alignment=al(1))
    fill_yellow(ws,9,18,9,31); merge(ws,'R9','AE9')
    merge(ws,'A10','B10','出生  ',alignment=al(1))
    fill_yellow(ws,10,3,10,11); merge(ws,'C10','K10')
    merge(ws,'P10','Q10','出生  ',alignment=al(1))
    fill_yellow(ws,10,18,10,25); merge(ws,'R10','Y10')
    merge(ws,'A11','B12','死亡  ',alignment=al(1))
    fill_yellow(ws,11,3,12,11); merge(ws,'C11','K12')
    fill_yellow(ws,11,16,12,18); merge(ws,'P11','R12','（　　）',alignment=al(1))
    fill_yellow(ws,11,25,12,28); merge(ws,'Y11','AB12')
    fill_yellow(ws,13,1,14,5); merge(ws,'A13','E14','（被相続人）',alignment=al(1))
    fill_yellow(ws,13,16,14,24); merge(ws,'P13','X14')
    fill_yellow(ws,15,1,15,9); merge(ws,'A15','I15')
    merge(ws,'P15','Q15')
    fill_yellow(ws,15,18,15,27); merge(ws,'R15','AA15')

    # 子1
    merge(ws,'P16','Q16','住所',alignment=al(1))
    fill_yellow(ws,16,18,16,31); merge(ws,'R16','AE16')
    merge(ws,'P17','Q18','出生',alignment=al(1))
    fill_yellow(ws,17,18,18,25); merge(ws,'R17','Y18')
    fill_yellow(ws,19,16,20,18); merge(ws,'P19','R20','（　　）',alignment=al(1))
    fill_yellow(ws,21,1,22,2); merge(ws,'A21','B22','住所　',alignment=al(1))
    fill_yellow(ws,21,3,22,12); merge(ws,'C21','L22')
    fill_yellow(ws,21,16,22,24); merge(ws,'P21','X22')
    merge(ws,'A23','B23','出生  ',alignment=al(0))
    fill_yellow(ws,23,3,23,11); merge(ws,'C23','K23')

    # 子2
    fill_yellow(ws,24,1,25,5); merge(ws,'A24','E25','（　　）',alignment=al(1))
    merge(ws,'P24','Q25','住所',alignment=al(1))
    fill_yellow(ws,24,18,25,31); merge(ws,'R24','AE25')
    fill_yellow(ws,26,1,26,9); merge(ws,'A26','I26')
    merge(ws,'P26','Q26','出生',alignment=al(1))
    fill_yellow(ws,26,18,26,25); merge(ws,'R26','Y26')
    fill_yellow(ws,27,16,27,18); merge(ws,'P27','R27','（　　）',alignment=al(1))
    fill_yellow(ws,28,16,29,24); merge(ws,'P28','X29')

    # 子3
    merge(ws,'P31','Q31','住所',alignment=al(1))
    fill_yellow(ws,31,18,31,31); merge(ws,'R31','AE31')
    merge(ws,'P32','Q32','出生',alignment=al(1))
    fill_yellow(ws,32,18,32,25); merge(ws,'R32','Y32')
    fill_yellow(ws,33,16,33,18); merge(ws,'P33','R33','（　　）',alignment=al(1))
    fill_yellow(ws,34,16,35,24); merge(ws,'P34','X35')

    # 子4（_child_start(4)=37 → ここではstatic=37に対応）
    s4 = _child_start(4)  # = 37
    _add_child_block_right(ws, s4)

    # 子5以降
    for n in range(5, num_children + 1):
        s = _child_start(n)
        _add_child_block_right(ws, s)

    # 以下余白
    merge(ws, f'P{wr}', f'T{wr}', '以下余白', alignment=al(2))

    # 作成者
    cr = wr + 3
    set_cell(ws,cr,11,'作成日：',alignment=al(0))
    fill_yellow(ws,cr,14,cr,24); merge(ws,f'N{cr}',f'X{cr}')
    set_cell(ws,cr+1,11,'作成者：',alignment=al(0))
    merge(ws,f'N{cr+1}',f'O{cr+1}','住所',alignment=al(1))
    fill_yellow(ws,cr+1,16,cr+1,28); merge(ws,f'P{cr+1}',f'AB{cr+1}')
    set_cell(ws,cr+2,14,'氏名',alignment=al(2))
    fill_yellow(ws,cr+2,16,cr+2,22); merge(ws,f'P{cr+2}',f'V{cr+2}')
    draw_creator_box(ws, cr-1)

    # 罫線（_sheet4と同じベース罫線）
    set_border(ws,13,13,bottom=1); set_border(ws,13,14,bottom=1); set_border(ws,13,15,bottom=1)
    set_border(ws,14,13,top=1,left=1); set_border(ws,14,14,top=1); set_border(ws,14,15,top=1)
    set_border(ws,15,13,left=1)
    set_border(ws,16,3,left=6); set_border(ws,16,13,left=1)
    set_border(ws,17,3,left=6); set_border(ws,17,13,left=1)
    set_border(ws,18,3,top=1,left=6)
    for c in range(4,13): set_border(ws,18,c,top=1)
    set_border(ws,18,13,left=1)
    set_border(ws,19,3,left=6); set_border(ws,19,13,left=1)
    set_border(ws,20,3,left=6); set_border(ws,20,13,left=1)
    set_border(ws,21,12,right=1); set_border(ws,21,13,left=1)
    set_border(ws,22,12,right=1)
    set_border(ws,22,13,top=1,left=1); set_border(ws,22,14,top=1); set_border(ws,22,15,top=1)
    for r in range(23,28): set_border(ws,r,13,left=1)
    set_border(ws,28,13,bottom=1,left=1); set_border(ws,28,14,bottom=1); set_border(ws,28,15,bottom=1)
    set_border(ws,29,13,top=1,left=1); set_border(ws,29,14,top=1); set_border(ws,29,15,top=1)
    for r in range(30,34): set_border(ws,r,13,left=1)
    set_border(ws,34,13,bottom=1,left=1); set_border(ws,34,14,bottom=1); set_border(ws,34,15,bottom=1)
    # 子4以降のM列罫線
    for n in range(4, num_children + 1):
        _add_m_borders_for_child(ws, _child_start(n))
    for c in range(16,25): set_border(ws,cr-1,c,bottom=1)

    # 申出人ラベルを選択した相続人の氏名欄右に配置
    merge(ws, appl_start, appl_end, '(申出人)', alignment=al(2))
