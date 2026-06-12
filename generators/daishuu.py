"""代襲相続が生じている場合（子の代襲）。

学習資料２の参考書式（記載例）を忠実に再現するテンプレート方式。
入力（配偶者の有無・孫の人数）に応じて最も近い公式書式を出力する。
"""
from openpyxl import Workbook
from openpyxl.styles import Font
from .common import (F11, F14, YEL, al, set_col_widths, set_row_heights,
                     merge, fill_yellow, set_border, set_cell)

F12 = Font(name="ＭＳ Ｐゴシック", size=12)
F10 = Font(name="ＭＳ Ｐゴシック", size=10)
_FONTS = {10: F10, 11: F11, 12: F12, 14: F14}


# ── 参考書式ブループリント（学習資料２より自動抽出）──────────
# 各要素: rh=行高(twips), merges=[start,end,値,黄塗り,横位置,字size],
#         singles=[セル,値,横位置,size], yfills=[黄塗りセル],
#         borders=[[セル,{辺:1=細,6=二重}]]
_BP = {'a': {'rh': {'4': 585, '5': 90, '6': 300, '7': 300, '8': 300, '9': 300, '10': 300, '11': 300, '12': 300, '13': 300, '14': 150, '15': 150, '16': 150, '17': 150, '18': 150, '19': 150, '20': 150, '21': 150, '22': 150, '23': 150, '24': 150, '25': 300, '26': 300, '28': 150, '29': 300, '30': 300, '31': 300, '32': 300, '34': 300, '35': 300, '36': 300, '37': 300, '38': 300, '39': 300, '40': 300, '41': 300}, 'merges': [['I4', 'L4', '被相続人', False, 3, 14], ['M4', 'S4', '', True, 2, 14], ['T4', 'Z4', '法定相続情報', False, 2, 14], ['A7', 'E7', '最後の住所', False, 1, 11], ['A8', 'K8', '', True, 2, 11], ['A9', 'E9', '最後の本籍', False, 1, 11], ['X10', 'AF10', '', True, 1, 11], ['A10', 'K10', '', True, 1, 11], ['A11', 'B11', '出生', False, 1, 11], ['X11', 'AE11', '', True, 1, 11], ['C11', 'J11', '', True, 1, 11], ['V11', 'W11', '出生  ', False, 1, 11], ['A12', 'B12', '死亡', False, 1, 11], ['C12', 'J12', '', True, 1, 11], ['V12', 'AB13', '（孫・代襲者）', False, 1, 11], ['A13', 'E13', '（被相続人）', False, 2, 11], ['V14', 'AB15', '', True, 1, 11], ['N14', 'S15', '被代襲者', False, 1, 11], ['A14', 'I15', '', True, 2, 11], ['M16', 'U17', '（\u3000\u3000\u3000\u3000\u3000\u3000\u3000\u3000\u3000\u3000\u3000\u3000）', True, 2, 11], ['U26', 'Y26', '以下余白', False, 2, 11], ['M29', 'W29', '', True, 2, 11], ['O30', 'Y30', '', True, 1, 11], ['M30', 'N30', '住所', False, 1, 11], ['O31', 'U31', '', True, 2, 11]], 'singles': [['V10', '住所', 0, 11], ['K29', '作成日：', 0, 11], ['K30', '作成者：', 0, 11], ['M31', '氏名', 2, 11]], 'yfills': [], 'borders': [['M14', {'bottom': 1}], ['J15', {'top': 1}], ['K15', {'top': 1}], ['L15', {'top': 1}], ['T15', {'top': 1}], ['U15', {'top': 1}], ['J28', {'top': 1, 'left': 1}], ['K28', {'top': 1}], ['L28', {'top': 1}], ['M28', {'top': 1}], ['N28', {'top': 1}], ['O28', {'top': 1}], ['P28', {'top': 1}], ['Q28', {'top': 1}], ['R28', {'top': 1}], ['S28', {'top': 1}], ['T28', {'top': 1}], ['U28', {'top': 1}], ['V28', {'top': 1}], ['W28', {'top': 1}], ['X28', {'top': 1}], ['Y28', {'top': 1}], ['Z28', {'top': 1}], ['AA28', {'top': 1, 'right': 1}], ['J29', {'left': 1}], ['AA29', {'right': 1}], ['J30', {'left': 1}], ['AA30', {'right': 1}], ['J31', {'left': 1}], ['AA31', {'right': 1}], ['J32', {'bottom': 1, 'left': 1}], ['K32', {'bottom': 1}], ['L32', {'bottom': 1}], ['M32', {'bottom': 1}], ['N32', {'bottom': 1}], ['O32', {'bottom': 1}], ['P32', {'bottom': 1}], ['Q32', {'bottom': 1}], ['R32', {'bottom': 1}], ['S32', {'bottom': 1}], ['T32', {'bottom': 1}], ['U32', {'bottom': 1}], ['V32', {'bottom': 1}], ['W32', {'bottom': 1}], ['X32', {'bottom': 1}], ['Y32', {'bottom': 1}], ['Z32', {'bottom': 1}], ['AA32', {'bottom': 1, 'right': 1}]]}, 'b2': {'rh': {'4': 585, '5': 90, '6': 300, '7': 300, '8': 300, '9': 300, '10': 300, '11': 150, '12': 150, '13': 150, '14': 150, '15': 300, '16': 300, '17': 300, '18': 150, '19': 150, '20': 150, '21': 150, '22': 150, '23': 150, '24': 300, '25': 150, '26': 150, '27': 300, '28': 300, '29': 150, '30': 150, '31': 150, '32': 150, '33': 300, '34': 300, '35': 300, '36': 300, '37': 300, '38': 300, '39': 300, '41': 150, '42': 300, '43': 300, '44': 300, '45': 300, '47': 300, '48': 300, '49': 300, '50': 300, '51': 300, '52': 300, '53': 300, '54': 300}, 'merges': [['N4', 'T4', '', True, 2, 14], ['U4', 'AA4', '法定相続情報', False, 2, 14], ['I4', 'M4', '被相続人', False, 3, 14], ['A7', 'E7', '最後の住所', False, 1, 11], ['A8', 'K8', '', True, 2, 11], ['O8', 'W9', '', True, 1, 11], ['M9', 'N9', '住所', False, 1, 11], ['A9', 'E9', '最後の本籍', False, 1, 11], ['O10', 'V10', '', True, 1, 11], ['M10', 'N10', '出生  ', False, 1, 11], ['A10', 'K10', '', True, 1, 11], ['C11', 'J12', '', True, 1, 11], ['M11', 'O12', '（\u3000）', True, 1, 11], ['A11', 'B12', '出生  ', False, 1, 11], ['A13', 'B14', '死亡  ', False, 1, 11], ['M13', 'S14', '', True, 1, 11], ['C13', 'J14', '', True, 1, 11], ['T13', 'W14', '(申出人)', False, 2, 11], ['A15', 'E15', '（被相続人）', False, 2, 11], ['X16', 'AF17', '', True, 1, 11], ['A16', 'I16', '', True, 1, 11], ['V17', 'W17', '住所', False, 1, 11], ['X18', 'AE19', '', True, 1, 11], ['V18', 'W19', '出生  ', False, 1, 11], ['C20', 'K23', '', True, 1, 11], ['V20', 'AB21', '（孫・代襲者）', False, 1, 11], ['A22', 'B23', '住所\u3000', False, 1, 11], ['V22', 'AB23', '', True, 1, 11], ['A24', 'B24', '出生  ', False, 0, 11], ['C24', 'J24', '', True, 1, 11], ['A25', 'D26', '（\u3000\u3000\u3000）', True, 1, 11], ['X25', 'AF27', '', True, 1, 11], ['M25', 'R26', '被代襲者', False, 1, 11], ['A27', 'I27', '', True, 1, 11], ['L27', 'T27', '（\u3000\u3000\u3000\u3000\u3000\u3000\u3000\u3000\u3000\u3000\u3000\u3000）', True, 1, 11], ['V28', 'W28', '出生', False, 1, 11], ['X28', 'AE28', '', True, 1, 11], ['V29', 'AB30', '（孫・代襲者）', False, 1, 11], ['V31', 'AB32', '', True, 1, 11], ['V39', 'Z39', '以下余白', False, 2, 11], ['N42', 'X42', '', True, 2, 11], ['N43', 'O43', '住所', False, 1, 11], ['P43', 'Z43', '', True, 1, 11], ['P44', 'V44', '', True, 2, 11]], 'singles': [['V27', '住所', 0, 11], ['K42', '作成日：', 0, 11], ['K43', '作成者：', 0, 11], ['N44', '氏名', 2, 11]], 'yfills': [], 'borders': [['L14', {'top': 1, 'left': 1}], ['L15', {'left': 1}], ['L16', {'left': 1}], ['C17', {'left': 6}], ['L17', {'left': 1}], ['C18', {'left': 6}], ['L18', {'left': 1}], ['C19', {'top': 1, 'left': 6}], ['D19', {'top': 1}], ['E19', {'top': 1}], ['F19', {'top': 1}], ['G19', {'top': 1}], ['H19', {'top': 1}], ['I19', {'top': 1}], ['J19', {'top': 1}], ['K19', {'top': 1}], ['L19', {'left': 1}], ['B20', {'right': 6}], ['C20', {'right': 1}], ['K20', {'right': 1}], ['L20', {'left': 1}], ['B21', {'right': 6}], ['K21', {'right': 1}], ['L21', {'left': 1}], ['K22', {'right': 1}], ['L22', {'left': 1}], ['K23', {'right': 1}], ['L23', {'left': 1}], ['U23', {'top': 1, 'left': 1}], ['L24', {'left': 1}], ['U24', {'left': 1}], ['L25', {'bottom': 1, 'left': 1}], ['U25', {'left': 1}], ['S26', {'top': 1}], ['T26', {'top': 1}], ['U26', {'left': 1}], ['L27', {'right': 1}], ['T27', {'right': 1}], ['U27', {'left': 1}], ['U28', {'left': 1}], ['U29', {'left': 1}], ['U30', {'left': 1}], ['U31', {'bottom': 1, 'left': 1}], ['J41', {'top': 1, 'left': 1}], ['K41', {'top': 1}], ['L41', {'top': 1}], ['M41', {'top': 1}], ['N41', {'top': 1}], ['O41', {'top': 1}], ['P41', {'top': 1}], ['Q41', {'top': 1}], ['R41', {'top': 1}], ['S41', {'top': 1}], ['T41', {'top': 1}], ['U41', {'top': 1}], ['V41', {'top': 1}], ['W41', {'top': 1}], ['X41', {'top': 1}], ['Y41', {'top': 1}], ['Z41', {'top': 1}], ['AA41', {'top': 1}], ['AB41', {'top': 1, 'right': 1}], ['J42', {'left': 1}], ['AB42', {'right': 1}], ['J43', {'left': 1}], ['AB43', {'right': 1}], ['J44', {'left': 1}], ['AB44', {'right': 1}], ['J45', {'bottom': 1, 'left': 1}], ['K45', {'bottom': 1}], ['L45', {'bottom': 1}], ['M45', {'bottom': 1}], ['N45', {'bottom': 1}], ['O45', {'bottom': 1}], ['P45', {'bottom': 1}], ['Q45', {'bottom': 1}], ['R45', {'bottom': 1}], ['S45', {'bottom': 1}], ['T45', {'bottom': 1}], ['U45', {'bottom': 1}], ['V45', {'bottom': 1}], ['W45', {'bottom': 1}], ['X45', {'bottom': 1}], ['Y45', {'bottom': 1}], ['Z45', {'bottom': 1}], ['AA45', {'bottom': 1}], ['AB45', {'bottom': 1, 'right': 1}]]}, 'b3': {'rh': {'4': 585, '5': 90, '6': 300, '7': 300, '8': 300, '9': 300, '10': 300, '11': 150, '12': 150, '13': 150, '14': 150, '15': 300, '16': 300, '17': 300, '18': 150, '19': 150, '20': 150, '21': 150, '22': 150, '23': 150, '24': 300, '25': 150, '26': 150, '27': 300, '28': 300, '29': 150, '30': 150, '31': 150, '32': 150, '33': 300, '34': 300, '35': 300, '36': 300, '37': 150, '38': 150, '39': 150, '40': 150, '41': 300, '42': 300, '44': 150, '45': 300, '46': 300, '47': 300, '48': 300, '50': 300, '51': 300, '52': 300, '53': 300, '54': 300, '55': 300, '56': 300, '57': 300}, 'merges': [['N4', 'T4', '', True, 2, 14], ['U4', 'AA4', '法定相続情報', False, 2, 14], ['I4', 'M4', '被相続人', False, 3, 14], ['A7', 'E7', '最後の住所', False, 1, 11], ['O8', 'W9', '', True, 1, 11], ['A8', 'K8', '', True, 2, 11], ['M9', 'N9', '住所', False, 1, 11], ['A9', 'E9', '最後の本籍', False, 1, 11], ['O10', 'V10', '', True, 1, 11], ['M10', 'N10', '出生  ', False, 1, 11], ['A10', 'K10', '', True, 1, 11], ['C11', 'J12', '', True, 1, 11], ['M11', 'O12', '（\u3000）', True, 1, 11], ['A11', 'B12', '出生  ', False, 1, 11], ['A13', 'B14', '死亡  ', False, 1, 11], ['M13', 'S14', '', True, 1, 11], ['C13', 'J14', '', True, 1, 11], ['T13', 'W14', '', False, 2, 11], ['A15', 'E15', '（被相続人）', False, 2, 11], ['X16', 'AF17', '', True, 1, 11], ['A16', 'I16', '', True, 1, 11], ['V17', 'W17', '住所', False, 1, 11], ['X18', 'AE19', '', True, 1, 11], ['V18', 'W19', '出生  ', False, 1, 11], ['C20', 'K23', '', True, 1, 11], ['V20', 'AB21', '（孫・代襲者）', False, 1, 11], ['A22', 'B23', '住所\u3000', False, 1, 11], ['V22', 'AB23', '', True, 1, 11], ['A24', 'B24', '出生  ', False, 0, 11], ['C24', 'J24', '', True, 1, 11], ['A25', 'D26', '（\u3000\u3000\u3000）', True, 1, 11], ['X25', 'AF27', '', True, 1, 11], ['M25', 'R26', '被代襲者', False, 1, 11], ['A27', 'I27', '', True, 1, 11], ['L27', 'T27', '（\u3000\u3000\u3000\u3000\u3000\u3000\u3000\u3000\u3000\u3000\u3000\u3000）', True, 1, 11], ['V28', 'W28', '出生', False, 1, 11], ['X28', 'AE28', '', True, 1, 11], ['V29', 'AB30', '（孫・代襲者）', False, 1, 11], ['V31', 'AB32', '', True, 1, 11], ['X34', 'AF35', '', True, 1, 11], ['V35', 'W35', '住所', False, 1, 11], ['V37', 'AB38', '（孫・代襲者）', False, 1, 11], ['V39', 'AB40', '', True, 1, 11], ['V42', 'Z42', '以下余白', False, 2, 11], ['N45', 'X45', '', True, 2, 11], ['N46', 'O46', '住所', False, 1, 11], ['P46', 'Z46', '', True, 1, 11], ['P47', 'V47', '', True, 2, 11]], 'singles': [['V27', '住所', 0, 11], ['V36', '出生  ', 0, 11], ['K45', '作成日：', 0, 11], ['K46', '作成者：', 0, 11], ['N47', '氏名', 2, 11]], 'yfills': ['X36', 'Y36', 'Z36', 'AA36', 'AB36', 'AC36', 'AD36', 'AE36'], 'borders': [['L14', {'top': 1, 'left': 1}], ['L15', {'left': 1}], ['L16', {'left': 1}], ['C17', {'left': 6}], ['L17', {'left': 1}], ['C18', {'left': 6}], ['L18', {'left': 1}], ['C19', {'top': 1, 'left': 6}], ['D19', {'top': 1}], ['E19', {'top': 1}], ['F19', {'top': 1}], ['G19', {'top': 1}], ['H19', {'top': 1}], ['I19', {'top': 1}], ['J19', {'top': 1}], ['K19', {'top': 1}], ['L19', {'left': 1}], ['B20', {'right': 6}], ['C20', {'right': 1}], ['K20', {'right': 1}], ['L20', {'left': 1}], ['B21', {'right': 6}], ['K21', {'right': 1}], ['L21', {'left': 1}], ['K22', {'right': 1}], ['L22', {'left': 1}], ['K23', {'right': 1}], ['L23', {'left': 1}], ['U23', {'top': 1, 'left': 1}], ['L24', {'left': 1}], ['U24', {'left': 1}], ['L25', {'bottom': 1, 'left': 1}], ['U25', {'left': 1}], ['S26', {'top': 1}], ['T26', {'top': 1}], ['U26', {'left': 1}], ['L27', {'right': 1}], ['T27', {'right': 1}], ['U27', {'left': 1}], ['U28', {'left': 1}], ['U29', {'left': 1}], ['U30', {'left': 1}], ['U31', {'bottom': 1, 'left': 1}], ['U32', {'top': 1, 'left': 1}], ['U33', {'left': 1}], ['U34', {'left': 1}], ['U35', {'left': 1}], ['U36', {'left': 1}], ['U37', {'left': 1}], ['U38', {'left': 1}], ['U39', {'bottom': 1, 'left': 1}], ['U40', {'top': 1}], ['J44', {'top': 1, 'left': 1}], ['K44', {'top': 1}], ['L44', {'top': 1}], ['M44', {'top': 1}], ['N44', {'top': 1}], ['O44', {'top': 1}], ['P44', {'top': 1}], ['Q44', {'top': 1}], ['R44', {'top': 1}], ['S44', {'top': 1}], ['T44', {'top': 1}], ['U44', {'top': 1}], ['V44', {'top': 1}], ['W44', {'top': 1}], ['X44', {'top': 1}], ['Y44', {'top': 1}], ['Z44', {'top': 1}], ['AA44', {'top': 1}], ['AB44', {'top': 1, 'right': 1}], ['J45', {'left': 1}], ['AB45', {'right': 1}], ['J46', {'left': 1}], ['AB46', {'right': 1}], ['J47', {'left': 1}], ['AB47', {'right': 1}], ['J48', {'bottom': 1, 'left': 1}], ['K48', {'bottom': 1}], ['L48', {'bottom': 1}], ['M48', {'bottom': 1}], ['N48', {'bottom': 1}], ['O48', {'bottom': 1}], ['P48', {'bottom': 1}], ['Q48', {'bottom': 1}], ['R48', {'bottom': 1}], ['S48', {'bottom': 1}], ['T48', {'bottom': 1}], ['U48', {'bottom': 1}], ['V48', {'bottom': 1}], ['W48', {'bottom': 1}], ['X48', {'bottom': 1}], ['Y48', {'bottom': 1}], ['Z48', {'bottom': 1}], ['AA48', {'bottom': 1}], ['AB48', {'bottom': 1, 'right': 1}]]}, 'b4': {'rh': {'4': 585, '5': 90, '6': 300, '7': 300, '8': 300, '9': 300, '10': 300, '11': 150, '12': 150, '13': 150, '14': 150, '15': 300, '16': 300, '17': 300, '18': 150, '19': 150, '20': 150, '21': 150, '22': 150, '23': 150, '24': 300, '25': 150, '26': 150, '27': 300, '28': 300, '29': 150, '30': 150, '31': 150, '32': 150, '33': 300, '34': 300, '35': 300, '36': 300, '37': 150, '38': 150, '39': 150, '40': 150, '41': 300, '42': 300, '44': 150, '45': 300, '46': 300, '47': 300, '48': 300, '50': 300, '51': 300, '52': 300, '53': 300, '54': 300, '55': 300, '56': 300, '57': 300}, 'merges': [['N4', 'T4', '', True, 2, 14], ['U4', 'AA4', '法定相続情報', False, 2, 14], ['I4', 'M4', '被相続人', False, 3, 14], ['A7', 'E7', '最後の住所', False, 1, 11], ['X7', 'AF8', '', True, 1, 11], ['V8', 'W8', '住所', False, 1, 11], ['A8', 'K8', '', True, 2, 11], ['X9', 'AE10', '', True, 1, 11], ['A9', 'E9', '最後の本籍', False, 1, 11], ['V9', 'W10', '出生  ', False, 1, 11], ['A10', 'K10', '', True, 1, 11], ['C11', 'J12', '', True, 1, 11], ['V11', 'AB12', '（孫・代襲者）', False, 1, 11], ['A11', 'B12', '出生  ', False, 1, 11], ['V13', 'AB14', '', True, 1, 11], ['A13', 'B14', '死亡  ', False, 1, 11], ['C13', 'J14', '', True, 1, 11], ['M13', 'R14', '被代襲者', False, 1, 11], ['A15', 'E15', '（被相続人）', False, 2, 11], ['M15', 'U15', '（\u3000\u3000\u3000\u3000\u3000\u3000\u3000\u3000\u3000\u3000\u3000\u3000）', True, 1, 11], ['X16', 'AF17', '', True, 1, 11], ['A16', 'I16', '', True, 1, 11], ['V17', 'W17', '住所', False, 1, 11], ['X18', 'AE19', '', True, 1, 11], ['V18', 'W19', '出生  ', False, 1, 11], ['V20', 'AB21', '（孫・代襲者）', False, 1, 11], ['C20', 'K23', '', True, 1, 11], ['A22', 'B23', '住所\u3000', False, 1, 11], ['V22', 'AB23', '', True, 1, 11], ['A24', 'B24', '出生  ', False, 0, 11], ['C24', 'J24', '', True, 1, 11], ['A25', 'D26', '（\u3000\u3000\u3000）', True, 1, 11], ['X25', 'AF27', '', True, 1, 11], ['M25', 'R26', '被代襲者', False, 1, 11], ['A27', 'I27', '', True, 1, 11], ['L27', 'T27', '（\u3000\u3000\u3000\u3000\u3000\u3000\u3000\u3000\u3000\u3000\u3000\u3000）', True, 1, 11], ['V28', 'W28', '出生', False, 1, 11], ['X28', 'AE28', '', True, 1, 11], ['V29', 'AB30', '（孫・代襲者）', False, 1, 11], ['V31', 'AB32', '', True, 1, 11], ['X34', 'AF35', '', True, 1, 11], ['V35', 'W35', '住所', False, 1, 11], ['V37', 'AB38', '（孫・代襲者）', False, 1, 11], ['V39', 'AB40', '', True, 1, 11], ['V42', 'Z42', '以下余白', False, 2, 11], ['N45', 'X45', '', True, 2, 11], ['N46', 'O46', '住所', False, 1, 11], ['P46', 'Z46', '', True, 1, 11], ['P47', 'V47', '', True, 2, 11]], 'singles': [['V27', '住所', 0, 11], ['V36', '出生  ', 0, 11], ['K45', '作成日：', 0, 11], ['K46', '作成者：', 0, 11], ['N47', '氏名', 2, 11]], 'yfills': ['X36', 'Y36', 'Z36', 'AA36', 'AB36', 'AC36', 'AD36', 'AE36'], 'borders': [['S13', {'bottom': 1}], ['T13', {'bottom': 1}], ['U13', {'bottom': 1}], ['L14', {'top': 1, 'left': 1}], ['L15', {'left': 1}], ['L16', {'left': 1}], ['C17', {'left': 6}], ['L17', {'left': 1}], ['C18', {'left': 6}], ['L18', {'left': 1}], ['C19', {'top': 1, 'left': 6}], ['D19', {'top': 1}], ['E19', {'top': 1}], ['F19', {'top': 1}], ['G19', {'top': 1}], ['H19', {'top': 1}], ['I19', {'top': 1}], ['J19', {'top': 1}], ['K19', {'top': 1}], ['L19', {'left': 1}], ['B20', {'right': 6}], ['C20', {'right': 1}], ['K20', {'right': 1}], ['L20', {'left': 1}], ['B21', {'right': 6}], ['K21', {'right': 1}], ['L21', {'left': 1}], ['K22', {'right': 1}], ['L22', {'left': 1}], ['K23', {'right': 1}], ['L23', {'left': 1}], ['U23', {'top': 1, 'left': 1}], ['L24', {'left': 1}], ['U24', {'left': 1}], ['L25', {'bottom': 1, 'left': 1}], ['U25', {'left': 1}], ['S26', {'top': 1}], ['T26', {'top': 1}], ['U26', {'left': 1}], ['L27', {'right': 1}], ['T27', {'right': 1}], ['U27', {'left': 1}], ['U28', {'left': 1}], ['U29', {'left': 1}], ['U30', {'left': 1}], ['U31', {'bottom': 1, 'left': 1}], ['U32', {'top': 1, 'left': 1}], ['U33', {'left': 1}], ['U34', {'left': 1}], ['U35', {'left': 1}], ['U36', {'left': 1}], ['U37', {'left': 1}], ['U38', {'left': 1}], ['U39', {'bottom': 1, 'left': 1}], ['U40', {'top': 1}], ['J44', {'top': 1, 'left': 1}], ['K44', {'top': 1}], ['L44', {'top': 1}], ['M44', {'top': 1}], ['N44', {'top': 1}], ['O44', {'top': 1}], ['P44', {'top': 1}], ['Q44', {'top': 1}], ['R44', {'top': 1}], ['S44', {'top': 1}], ['T44', {'top': 1}], ['U44', {'top': 1}], ['V44', {'top': 1}], ['W44', {'top': 1}], ['X44', {'top': 1}], ['Y44', {'top': 1}], ['Z44', {'top': 1}], ['AA44', {'top': 1}], ['AB44', {'top': 1, 'right': 1}], ['J45', {'left': 1}], ['AB45', {'right': 1}], ['J46', {'left': 1}], ['AB46', {'right': 1}], ['J47', {'left': 1}], ['AB47', {'right': 1}], ['J48', {'bottom': 1, 'left': 1}], ['K48', {'bottom': 1}], ['L48', {'bottom': 1}], ['M48', {'bottom': 1}], ['N48', {'bottom': 1}], ['O48', {'bottom': 1}], ['P48', {'bottom': 1}], ['Q48', {'bottom': 1}], ['R48', {'bottom': 1}], ['S48', {'bottom': 1}], ['T48', {'bottom': 1}], ['U48', {'bottom': 1}], ['V48', {'bottom': 1}], ['W48', {'bottom': 1}], ['X48', {'bottom': 1}], ['Y48', {'bottom': 1}], ['Z48', {'bottom': 1}], ['AA48', {'bottom': 1}], ['AB48', {'bottom': 1, 'right': 1}]]}, 'b5': {'rh': {'4': 585, '5': 90, '6': 300, '7': 300, '8': 300, '9': 300, '10': 150, '11': 150, '12': 150, '13': 150, '14': 150, '15': 150, '16': 300, '17': 300, '18': 300, '19': 300, '20': 300, '21': 150, '22': 150, '23': 150, '24': 150, '25': 300, '26': 300, '27': 300, '28': 150, '29': 150, '30': 150, '31': 150, '32': 150, '33': 150, '34': 300, '35': 150, '36': 150, '37': 300, '38': 300, '39': 150, '40': 150, '41': 150, '42': 150, '43': 300, '44': 300, '45': 300, '46': 300, '47': 150, '48': 150, '49': 150, '50': 150, '51': 300, '52': 300, '54': 150, '55': 300, '56': 300, '57': 300, '58': 300, '60': 300, '61': 300, '62': 300, '63': 300, '64': 300, '65': 300, '66': 300, '67': 300}, 'merges': [['N4', 'T4', '', True, 2, 14], ['U4', 'AA4', '法定相続情報', False, 2, 14], ['I4', 'M4', '被相続人', False, 3, 14], ['X8', 'AF9', '', True, 1, 11], ['V9', 'W9', '住所', False, 1, 11], ['X10', 'AE11', '', True, 1, 11], ['V10', 'W11', '出生  ', False, 1, 11], ['V12', 'AB13', '（孫・代襲者）', False, 1, 11], ['V14', 'AB15', '', True, 1, 11], ['X17', 'AF18', '', True, 1, 11], ['A17', 'E17', '最後の住所', False, 1, 11], ['A18', 'K18', '', True, 2, 11], ['V18', 'W18', '住所', False, 1, 11], ['A19', 'E19', '最後の本籍', False, 1, 11], ['V19', 'W20', '出生  ', False, 1, 11], ['X19', 'AE20', '', True, 1, 11], ['A20', 'K20', '', True, 1, 11], ['C21', 'J22', '', True, 1, 11], ['A21', 'B22', '出生  ', False, 1, 11], ['V21', 'AB22', '（孫・代襲者）', False, 1, 11], ['M23', 'R24', '被代襲者', False, 1, 11], ['V23', 'AB24', '', True, 1, 11], ['C23', 'J24', '', True, 1, 11], ['A23', 'B24', '死亡  ', False, 1, 11], ['M25', 'U25', '（\u3000\u3000\u3000\u3000\u3000\u3000\u3000\u3000\u3000\u3000\u3000\u3000）', True, 1, 11], ['A25', 'E25', '（被相続人）', False, 2, 11], ['X26', 'AF27', '', True, 1, 11], ['A26', 'I26', '', True, 1, 11], ['V27', 'W27', '住所', False, 1, 11], ['X28', 'AE29', '', True, 1, 11], ['V28', 'W29', '出生  ', False, 1, 11], ['V30', 'AB31', '（孫・代襲者）', False, 1, 11], ['C30', 'K33', '', True, 1, 11], ['A32', 'B33', '住所\u3000', False, 1, 11], ['V32', 'AB33', '', True, 1, 11], ['C34', 'J34', '', True, 1, 11], ['A34', 'B34', '出生  ', False, 0, 11], ['X35', 'AF37', '', True, 1, 11], ['M35', 'R36', '被代襲者', False, 1, 11], ['A35', 'D36', '（\u3000\u3000\u3000）', True, 1, 11], ['A37', 'I37', '', True, 1, 11], ['L37', 'T37', '（\u3000\u3000\u3000\u3000\u3000\u3000\u3000\u3000\u3000\u3000\u3000\u3000）', True, 1, 11], ['V38', 'W38', '出生', False, 1, 11], ['X38', 'AE38', '', True, 1, 11], ['V39', 'AB40', '（孫・代襲者）', False, 1, 11], ['V41', 'AB42', '', True, 1, 11], ['X44', 'AF45', '', True, 1, 11], ['V45', 'W45', '住所', False, 1, 11], ['V47', 'AB48', '（孫・代襲者）', False, 1, 11], ['V49', 'AB50', '', True, 1, 11], ['V52', 'Z52', '以下余白', False, 2, 11], ['N55', 'X55', '', True, 2, 11], ['N56', 'O56', '住所', False, 1, 11], ['P56', 'Z56', '', True, 1, 11], ['P57', 'V57', '', True, 2, 11]], 'singles': [['V37', '住所', 0, 11], ['V46', '出生  ', 0, 11], ['K55', '作成日：', 0, 11], ['K56', '作成者：', 0, 11], ['N57', '氏名', 2, 11]], 'yfills': ['X46', 'Y46', 'Z46', 'AA46', 'AB46', 'AC46', 'AD46', 'AE46'], 'borders': [['U15', {'top': 1, 'left': 1}], ['U16', {'left': 1}], ['U17', {'left': 1}], ['U18', {'left': 1}], ['U19', {'left': 1}], ['U20', {'left': 1}], ['U21', {'left': 1}], ['U22', {'left': 1}], ['S23', {'bottom': 1}], ['T23', {'bottom': 1}], ['U23', {'bottom': 1, 'left': 1}], ['L24', {'top': 1, 'left': 1}], ['L25', {'left': 1}], ['L26', {'left': 1}], ['C27', {'left': 6}], ['L27', {'left': 1}], ['C28', {'left': 6}], ['L28', {'left': 1}], ['C29', {'top': 1, 'left': 6}], ['D29', {'top': 1}], ['E29', {'top': 1}], ['F29', {'top': 1}], ['G29', {'top': 1}], ['H29', {'top': 1}], ['I29', {'top': 1}], ['J29', {'top': 1}], ['K29', {'top': 1}], ['L29', {'left': 1}], ['B30', {'right': 6}], ['C30', {'right': 1}], ['K30', {'right': 1}], ['L30', {'left': 1}], ['B31', {'right': 6}], ['K31', {'right': 1}], ['L31', {'left': 1}], ['K32', {'right': 1}], ['L32', {'left': 1}], ['K33', {'right': 1}], ['L33', {'left': 1}], ['U33', {'top': 1, 'left': 1}], ['L34', {'left': 1}], ['U34', {'left': 1}], ['L35', {'bottom': 1, 'left': 1}], ['U35', {'left': 1}], ['S36', {'top': 1}], ['T36', {'top': 1}], ['U36', {'left': 1}], ['L37', {'right': 1}], ['T37', {'right': 1}], ['U37', {'left': 1}], ['U38', {'left': 1}], ['U39', {'left': 1}], ['U40', {'left': 1}], ['U41', {'bottom': 1, 'left': 1}], ['U42', {'top': 1, 'left': 1}], ['U43', {'left': 1}], ['U44', {'left': 1}], ['U45', {'left': 1}], ['U46', {'left': 1}], ['U47', {'left': 1}], ['U48', {'left': 1}], ['U49', {'bottom': 1, 'left': 1}], ['U50', {'top': 1}], ['J54', {'top': 1, 'left': 1}], ['K54', {'top': 1}], ['L54', {'top': 1}], ['M54', {'top': 1}], ['N54', {'top': 1}], ['O54', {'top': 1}], ['P54', {'top': 1}], ['Q54', {'top': 1}], ['R54', {'top': 1}], ['S54', {'top': 1}], ['T54', {'top': 1}], ['U54', {'top': 1}], ['V54', {'top': 1}], ['W54', {'top': 1}], ['X54', {'top': 1}], ['Y54', {'top': 1}], ['Z54', {'top': 1}], ['AA54', {'top': 1}], ['AB54', {'top': 1, 'right': 1}], ['J55', {'left': 1}], ['AB55', {'right': 1}], ['J56', {'left': 1}], ['AB56', {'right': 1}], ['J57', {'left': 1}], ['AB57', {'right': 1}], ['J58', {'bottom': 1, 'left': 1}], ['K58', {'bottom': 1}], ['L58', {'bottom': 1}], ['M58', {'bottom': 1}], ['N58', {'bottom': 1}], ['O58', {'bottom': 1}], ['P58', {'bottom': 1}], ['Q58', {'bottom': 1}], ['R58', {'bottom': 1}], ['S58', {'bottom': 1}], ['T58', {'bottom': 1}], ['U58', {'bottom': 1}], ['V58', {'bottom': 1}], ['W58', {'bottom': 1}], ['X58', {'bottom': 1}], ['Y58', {'bottom': 1}], ['Z58', {'bottom': 1}], ['AA58', {'bottom': 1}], ['AB58', {'bottom': 1, 'right': 1}]]}}


def _apply(ws, bp):
    set_col_widths(ws)
    set_row_heights(ws, {int(k): v for k, v in bp["rh"].items()})
    # マージ（値・塗り・配置・フォント）
    for start, end, val, fill, h, sz in bp["merges"]:
        merge(ws, start, end,
              value=(val if val != "" else None),
              font=_FONTS.get(sz, F11),
              fill=(YEL if fill else None),
              alignment=al(h))
    # 単独セルの値
    for coord, val, h, sz in bp["singles"]:
        c = ws[coord]
        c.value = val
        c.font = _FONTS.get(sz, F11)
        c.alignment = al(h)
    # 単独セルの黄塗り
    for coord in bp["yfills"]:
        ws[coord].fill = YEL
    # 罫線
    from openpyxl.utils.cell import coordinate_to_tuple
    for coord, sides in bp["borders"]:
        r, cc = coordinate_to_tuple(coord)
        set_border(ws, r, cc, **sides)


def _sheet_name_tpl(key, total_gc):
    if key == 'a':
        return '子1人・孫1人の場合'
    return f'配偶者・子2人・孫{total_gc}人の場合'


def _normalize(children_data):
    kids = []
    for c in (children_data or []):
        alive = bool(c.get('alive', True))
        gc = 0 if alive else max(int(c.get('num_grandchildren', 0) or 0), 0)
        kids.append({'alive': alive, 'gc': gc})
    if not kids:
        kids = [{'alive': False, 'gc': 1}]
    return kids


def _match_template(has_spouse, kids):
    """入力が参考書式（学習資料２）と同一の関係構成なら該当キーを返す。

    参考書式の実際の構成：
      a  = 配偶者なし・子1人（死亡）・孫1人
      b2 = 配偶者あり・子2人（1人生存・1人死亡）・孫2人
      b3 = 配偶者あり・子2人（1人生存・1人死亡）・孫3人
      b4 = 配偶者あり・子2人（2人とも死亡）・孫1人＋3人
      b5 = 配偶者あり・子2人（2人とも死亡）・孫2人＋3人
    """
    dead = sorted(k['gc'] for k in kids if not k['alive'])
    alive = sum(1 for k in kids if k['alive'])
    if not has_spouse:
        if alive == 0 and dead == [1]:
            return 'a'
        return None
    if len(kids) == 2:
        if alive == 1 and dead in ([2], [3]):
            return f'b{dead[0]}'
        if alive == 0 and dead == [1, 3]:
            return 'b4'
        if alive == 0 and dead == [2, 3]:
            return 'b5'
    return None


def _dyn_name(has_spouse, kids):
    n = len(kids)
    gc = sum(k['gc'] for k in kids if not k['alive'])
    parts = []
    if has_spouse:
        parts.append('配偶者')
    parts.append(f'子{n}人')
    parts.append(f'孫{gc}人（代襲）')
    name = '・'.join(parts) + 'の場合'
    return name if len(name) <= 31 else '代襲相続の場合'


def _build_dynamic(ws, has_spouse, kids):
    """参考書式にない構成を、学習した配置文法から外挿して描画する。

    関係線の文法（学習資料２より解読）:
      婚姻線   : B/C列境界の二重縦線。被相続人氏名(A16:I16)の下から
                 配偶者欄(C20:K23)へ。
      子の幹線 : L列左端の細縦線。婚姻線の中間（行18/19境界）から
                 横線C-K列で接続し、各子のノード位置に小枝を出す。
      生存子   : センター上段(行8-14)または左ゾーンのブロック。
                 氏名欄の中央の高さで幹線に接続。
      被代襲者 : ラベル(M:R 2行)＋氏名(A:I)＋（…）状態欄(L:T)。
                 ラベル中央の高さで幹線に接続し、S-T列の横線から
                 U列の孫幹線へ。
      孫       : 右ゾーン(V-AF)の8行ブロック。氏名欄の中央の高さで
                 孫幹線に接続。
    """
    set_col_widths(ws)
    rh = {4: 585, 5: 90, 6: 300, 7: 300, 8: 300, 9: 300, 10: 300,
          11: 150, 12: 150, 13: 150, 14: 150, 15: 300, 16: 300}

    # ── ヘッダー・被相続人 ──
    merge(ws, 'I4', 'M4', '被相続人', F14, alignment=al(3))
    merge(ws, 'N4', 'T4', None, None, YEL)
    merge(ws, 'U4', 'AA4', '法定相続情報', F14, alignment=al(2))
    merge(ws, 'A7', 'E7', '最後の住所', alignment=al(1))
    merge(ws, 'A8', 'K8', None, None, YEL)
    merge(ws, 'A9', 'E9', '最後の本籍', alignment=al(1))
    merge(ws, 'A10', 'K10', None, None, YEL)
    merge(ws, 'A11', 'B12', '出生  ', alignment=al(1))
    merge(ws, 'C11', 'J12', None, None, YEL)
    merge(ws, 'A13', 'B14', '死亡  ', alignment=al(1))
    merge(ws, 'C13', 'J14', None, None, YEL)
    merge(ws, 'A15', 'E15', '（被相続人）', alignment=al(2))
    merge(ws, 'A16', 'I16', None, None, YEL)   # 被相続人氏名

    # ── 子の表示順：最初の生存子はセンター上段（b2の文法）──
    rest = list(kids)
    top_child = None
    for i, k in enumerate(rest):
        if k['alive']:
            top_child = rest.pop(i)
            break

    stub_gls = []          # 幹線に接続する子ノードの高さ（行境界）
    if top_child is not None:
        merge(ws, 'O8', 'W9', None, None, YEL)
        merge(ws, 'M9', 'N9', '住所', alignment=al(1))
        merge(ws, 'M10', 'N10', '出生  ', alignment=al(1))
        merge(ws, 'O10', 'V10', None, None, YEL)
        merge(ws, 'M11', 'O12', '（　）', None, YEL, al(1))
        merge(ws, 'M13', 'S14', None, None, YEL)
        merge(ws, 'T13', 'W14', '(申出人)', alignment=al(2))
        stub_gls.append(13)

    FEED_GL = 18           # 婚姻線→幹線の横線（行18/19境界）

    # ── 配偶者と婚姻線 ──
    if has_spouse:
        set_border(ws, 17, 3, left=6)
        set_border(ws, 18, 3, left=6)
        set_border(ws, 19, 3, top=1, left=6)
        set_border(ws, 20, 2, right=6)
        set_border(ws, 21, 2, right=6)
        merge(ws, 'C20', 'K23', None, None, YEL)
        for r in range(20, 24):
            set_border(ws, r, 11, right=1)
        merge(ws, 'A22', 'B23', '住所　', alignment=al(1))
        merge(ws, 'A24', 'B24', '出生  ', alignment=al(1))
        merge(ws, 'C24', 'J24', None, None, YEL)
        merge(ws, 'A25', 'D26', '（　　　）', None, YEL, al(1))
        rh.update({17: 300, 18: 150, 19: 150, 20: 150, 21: 150,
                   22: 150, 23: 150, 24: 300, 25: 150, 26: 150})
        left_cur = 28
    else:
        # 配偶者なし：被相続人氏名から細線で幹線へ
        set_border(ws, 17, 3, left=1)
        set_border(ws, 18, 3, left=1)
        set_border(ws, 19, 3, top=1)
        rh.update({17: 300, 18: 150, 19: 150})
        left_cur = 20

    # 幹線への横枝線（C-K列、行18/19境界）
    for c in range(4, 12):
        set_border(ws, 19, c, top=1)

    right_cur = 16         # 右ゾーン（孫ブロック）の開始行
    gc_specs = []          # (被代襲者の接続高さ, [孫氏名の接続高さ...])

    for k in rest:
        if k['alive']:
            # ── 生存子（左ゾーン）──
            r = left_cur
            merge(ws, f'A{r}', f'B{r+1}', '住所　', alignment=al(1))
            merge(ws, f'C{r}', f'K{r+1}', None, None, YEL)
            merge(ws, f'A{r+2}', f'B{r+2}', '出生  ', alignment=al(1))
            merge(ws, f'C{r+2}', f'J{r+2}', None, None, YEL)
            merge(ws, f'A{r+3}', f'I{r+4}', None, None, YEL)
            merge(ws, f'M{r+3}', f'O{r+4}', '（　）', None, YEL, al(1))
            stub_gls.append(r + 3)
            rh.update({r: 150, r + 1: 150, r + 2: 300,
                       r + 3: 300, r + 4: 300, r + 5: 150})
            left_cur = r + 6
        else:
            # ── 被代襲者（死亡した子）──
            r = left_cur
            merge(ws, f'M{r}', f'R{r+1}', '被代襲者', alignment=al(2))
            merge(ws, f'A{r+2}', f'I{r+2}', None, None, YEL)
            merge(ws, f'L{r+2}', f'T{r+2}',
                  '（　　　　　　　　　　　　）', None, YEL, al(1))
            stub_gls.append(r)
            rh.update({r: 150, r + 1: 150, r + 2: 300, r + 3: 150})
            # ── この子の孫グループ（右ゾーン）──
            names = []
            for _ in range(k['gc']):
                g = right_cur
                merge(ws, f'X{g}', f'AF{g+1}', None, None, YEL)
                merge(ws, f'V{g+1}', f'W{g+1}', '住所', alignment=al(1))
                merge(ws, f'V{g+2}', f'W{g+3}', '出生  ', alignment=al(1))
                merge(ws, f'X{g+2}', f'AE{g+3}', None, None, YEL)
                merge(ws, f'V{g+4}', f'AB{g+5}', '（孫・代襲者）',
                      alignment=al(1))
                merge(ws, f'V{g+6}', f'AB{g+7}', None, None, YEL)
                names.append(g + 6)
                rh.setdefault(g, 300)
                for rr in range(g + 1, g + 9):
                    rh.setdefault(rr, 150)
                right_cur = g + 9
            if names:
                gc_specs.append((r, names))
                # 被代襲者→孫幹線への横線（S-T列）と（…）右端の縦線
                set_border(ws, r + 1, 19, top=1)
                set_border(ws, r + 1, 20, top=1)
                set_border(ws, r + 2, 20, right=1)
            left_cur = r + 4

    # ── 子の幹線（L列左端）──
    lo = min(stub_gls + [FEED_GL])
    hi = max(stub_gls + [FEED_GL])
    for rr in range(lo + 1, hi + 1):
        set_border(ws, rr, 12, left=1)
    for g in stub_gls:
        set_border(ws, g + 1, 12, top=1)

    # ── 孫幹線（U列左端）──
    for fgl, names in gc_specs:
        lo = min([fgl] + names)
        hi = max([fgl] + names)
        for rr in range(lo + 1, hi + 1):
            set_border(ws, rr, 21, left=1)
        for g in names:
            set_border(ws, g + 1, 21, top=1)

    # ── 以下余白・作成者欄 ──
    e = max(left_cur, right_cur)
    merge(ws, f'V{e}', f'Z{e}', '以下余白', alignment=al(2))
    rh[e] = 300
    b = e + 2
    set_border(ws, b, 10, top=1, left=1)
    for c in range(11, 28):
        set_border(ws, b, c, top=1)
    set_border(ws, b, 28, top=1, right=1)
    for rr in range(b + 1, b + 4):
        set_border(ws, rr, 10, left=1)
        set_border(ws, rr, 28, right=1)
    set_border(ws, b + 4, 10, bottom=1, left=1)
    for c in range(11, 28):
        set_border(ws, b + 4, c, bottom=1)
    set_border(ws, b + 4, 28, bottom=1, right=1)
    set_cell(ws, b + 1, 11, '作成日：', alignment=al(0))
    merge(ws, f'N{b+1}', f'X{b+1}', None, None, YEL)
    set_cell(ws, b + 2, 11, '作成者：', alignment=al(0))
    merge(ws, f'N{b+2}', f'O{b+2}', '住所', alignment=al(1))
    merge(ws, f'P{b+2}', f'Z{b+2}', None, None, YEL)
    set_cell(ws, b + 3, 14, '氏名', alignment=al(2))
    merge(ws, f'P{b+3}', f'V{b+3}', None, None, YEL)
    rh[b] = 150
    for rr in range(b + 1, b + 5):
        rh[rr] = 300
    for rr in range(6, b + 5):
        rh.setdefault(rr, 150)
    set_row_heights(ws, rh)


def generate(has_spouse=True, children_data=None,
             appl_start=None, appl_end=None) -> Workbook:
    """子の代襲相続の書式を生成する。

    参考書式（学習資料２）と同一の関係構成はテンプレートを忠実に
    再現し、それ以外の構成は学習した配置文法から関係図を外挿して
    動的に生成する。
    """
    kids = _normalize(children_data)
    wb = Workbook()
    wb.remove(wb.active)

    key = _match_template(has_spouse, kids)
    if key:
        total_gc = sum(k['gc'] for k in kids if not k['alive'])
        ws = wb.create_sheet(_sheet_name_tpl(key, total_gc))
        _apply(ws, _BP[key])
        return wb

    ws = wb.create_sheet(_dyn_name(has_spouse, kids))
    _build_dynamic(ws, has_spouse, kids)
    return wb
