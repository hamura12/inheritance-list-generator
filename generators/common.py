"""共通ヘルパー関数"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

F14 = Font(name="ＭＳ Ｐゴシック", size=14)
F11 = Font(name="ＭＳ Ｐゴシック", size=11)
YEL = PatternFill(fgColor="FFFF00", fill_type="solid")
TH = Side(style='thin')
DB = Side(style='double')
NO = Side(style=None)


def al(h, v='center'):
    hmap = {0: 'general', 1: 'left', 2: 'center', 3: 'right'}
    return Alignment(horizontal=hmap[h], vertical=v)


def set_col_widths(ws, n=32):
    for i in range(1, n + 1):
        ws.column_dimensions[get_column_letter(i)].width = 2.5


def set_row_heights(ws, heights: dict):
    for row, twips in heights.items():
        ws.row_dimensions[row].height = twips / 20.0


def merge(ws, start, end, value=None, font=None, fill=None, alignment=None):
    from openpyxl.utils import range_boundaries
    min_col, min_row, max_col, max_row = range_boundaries(f"{start}:{end}")
    # 既存マージと重複する範囲を先に解除してから再マージ
    to_remove = []
    for mr in ws.merged_cells.ranges:
        b = range_boundaries(str(mr))
        if not (b[2] < min_col or b[0] > max_col or b[3] < min_row or b[1] > max_row):
            to_remove.append(str(mr))
    for r in to_remove:
        ws.unmerge_cells(r)
    ws.merge_cells(f"{start}:{end}")
    cell = ws[start]
    if value is not None:
        cell.value = value
        cell.font = font or F11
    if fill is not None:
        cell.fill = fill
    if alignment is not None:
        cell.alignment = alignment


def fill_yellow(ws, r1, c1, r2, c2):
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            ws.cell(r, c).fill = YEL


def set_border(ws, row, col, top=None, bottom=None, left=None, right=None):
    def side(n):
        if n is None:
            return None
        return TH if n == 1 else (DB if n == 6 else NO)

    cell = ws.cell(row, col)
    old = cell.border
    cell.border = Border(
        top=side(top) if top is not None else old.top,
        bottom=side(bottom) if bottom is not None else old.bottom,
        left=side(left) if left is not None else old.left,
        right=side(right) if right is not None else old.right,
    )


def set_cell(ws, row, col, value, font=None, alignment=None):
    cell = ws.cell(row, col)
    cell.value = value
    cell.font = font or F11
    if alignment:
        cell.alignment = alignment


def draw_creator_box(ws, top_row, left_col=10, right_col=28):
    br = top_row + 4
    set_border(ws, top_row, left_col, top=1, left=1)
    for c in range(left_col + 1, right_col):
        set_border(ws, top_row, c, top=1)
    set_border(ws, top_row, right_col, top=1, right=1)
    for r in range(top_row + 1, br):
        set_border(ws, r, left_col, left=1)
        set_border(ws, r, right_col, right=1)
    set_border(ws, br, left_col, bottom=1, left=1)
    for c in range(left_col + 1, right_col):
        set_border(ws, br, c, bottom=1)
    set_border(ws, br, right_col, bottom=1, right=1)
