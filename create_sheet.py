#!/usr/bin/env python3
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()
wb.remove(wb.active)

F14 = Font(name="ＭＳ Ｐゴシック", size=14)
F11 = Font(name="ＭＳ Ｐゴシック", size=11)
YEL = PatternFill(fgColor="FFFF00", fill_type="solid")
TH = Side(style='thin')
DB = Side(style='double')
NO = Side(style=None)

def al(h, v='center'):
    hmap = {0:'general', 1:'left', 2:'center', 3:'right'}
    return Alignment(horizontal=hmap[h], vertical=v)

def cw(ws, n=32):
    for i in range(1, n+1):
        ws.column_dimensions[get_column_letter(i)].width = 2.5

def rh(ws, d):
    for r, h in d.items():
        ws.row_dimensions[r].height = h / 20.0

def mg(ws, s, e, v=None, f=F11, fi=None, a=None):
    ws.merge_cells(f"{s}:{e}")
    c = ws[s]
    if v is not None: c.value = v; c.font = f
    if fi is not None: c.fill = fi
    if a is not None: c.alignment = a

def yf(ws, r1, c1, r2, c2):
    for r in range(r1, r2+1):
        for c in range(c1, c2+1):
            ws.cell(r, c).fill = YEL

def bd(ws, r, c, t=None, b=None, l=None, ri=None):
    def s(n):
        if n is None: return None
        return TH if n==1 else (DB if n==6 else NO)
    cell = ws.cell(r, c)
    old = cell.border
    cell.border = Border(
        top=s(t) if t is not None else old.top,
        bottom=s(b) if b is not None else old.bottom,
        left=s(l) if l is not None else old.left,
        right=s(ri) if ri is not None else old.right,
    )

def cv(ws, r, c, v, f=F11, a=None):
    cell = ws.cell(r, c)
    cell.value = v; cell.font = f
    if a: cell.alignment = a

def creator_box(ws, top_row, left_col=10, right_col=28):
    br = top_row + 4
    bd(ws, top_row, left_col, t=1, l=1)
    for c in range(left_col+1, right_col): bd(ws, top_row, c, t=1)
    bd(ws, top_row, right_col, t=1, ri=1)
    for r in range(top_row+1, br):
        bd(ws, r, left_col, l=1); bd(ws, r, right_col, ri=1)
    bd(ws, br, left_col, b=1, l=1)
    for c in range(left_col+1, right_col): bd(ws, br, c, b=1)
    bd(ws, br, right_col, b=1, ri=1)

# ─────────────────────────────────────────────
# SHEET 1: 配偶者・子1人の場合
# ─────────────────────────────────────────────
ws = wb.create_sheet('配偶者・子1人の場合')
cw(ws)
rh(ws, {4:585,5:90,6:300,7:300,8:300,9:300,10:300,11:150,12:150,
        13:300,14:300,15:300,16:150,17:150,18:150,19:150,20:300,21:300,
        22:150,23:150,24:300,26:150,27:300,28:300,29:300,30:300,
        32:300,33:300,34:300,35:300,36:300,37:300,38:300,39:300})

# Header row 4
mg(ws,'I4','M4','被相続人',F14,a=al(3))
yf(ws,4,14,4,20); mg(ws,'N4','T4')
mg(ws,'U4','AA4','法定相続情報',F14,a=al(2))

# 被相続人 info left side
mg(ws,'A6','E6','最後の住所',a=al(1))
yf(ws,7,1,7,14); mg(ws,'A7','N7')
mg(ws,'A8','E8','最後の本籍',a=al(1))
yf(ws,9,1,9,14); mg(ws,'A9','N9')
mg(ws,'P9','Q9')
yf(ws,9,18,9,27); mg(ws,'R9','AA9')
mg(ws,'A10','B10','出生  ',a=al(1))
yf(ws,10,3,10,11); mg(ws,'C10','K10')
mg(ws,'A11','B12','死亡  ',a=al(1))
yf(ws,11,3,12,11); mg(ws,'C11','K12')

# Row 13-14: 被相続人 label + 配偶者 start
mg(ws,'A13','E13','（被相続人）',a=al(2))
mg(ws,'P13','Q13','住所',a=al(1))
yf(ws,13,18,13,31); mg(ws,'R13','AE13')
yf(ws,14,1,14,9); mg(ws,'A14','I14')
mg(ws,'P14','Q14','出生  ',a=al(1))
yf(ws,14,18,14,27); mg(ws,'R14','AA14')

# Row 15: 配偶者 name
yf(ws,15,16,15,18); mg(ws,'P15','R15','（　　）',a=al(1))

# Rows 16-19: 配偶者 body + 申出人
yf(ws,16,16,17,24); mg(ws,'P16','X17')
mg(ws,'Y16','AB17','(申出人)',a=al(2))
yf(ws,18,16,19,24); mg(ws,'P18','X19')

# 子1 area rows 20-24 left + right
mg(ws,'A20','B20','住所　',a=al(1))
yf(ws,20,3,20,15); mg(ws,'C20','O20')
mg(ws,'P20','Q20')
yf(ws,20,18,20,27); mg(ws,'R20','AA20')
mg(ws,'A21','B21','出生  ',a=al(0))
yf(ws,21,3,21,11); mg(ws,'C21','K21')
yf(ws,22,1,23,5); mg(ws,'A22','E23','（　　）',a=al(1))
yf(ws,22,16,23,24); mg(ws,'P22','X23')
yf(ws,24,1,24,9); mg(ws,'A24','I24')
mg(ws,'P24','T24','以下余白',a=al(2))

# Creator rows 26-30
cv(ws,27,11,'作成日：',a=al(0))
yf(ws,27,14,27,24); mg(ws,'N27','X27')
cv(ws,28,11,'作成者：',a=al(0))
mg(ws,'N28','O28','住所',a=al(1))
yf(ws,28,16,28,28); mg(ws,'P28','AB28')
cv(ws,29,14,'氏名',a=al(2))
yf(ws,29,16,29,22); mg(ws,'P29','V29')

# Tree connecting lines
bd(ws,15,3,l=6); bd(ws,16,3,l=6)
bd(ws,16,13,b=1); bd(ws,16,14,b=1); bd(ws,16,15,b=1)
bd(ws,17,3,t=1,l=6)
for c in range(4,16): bd(ws,17,c,t=1)
bd(ws,18,3,l=6); bd(ws,19,3,l=6)

# Creator box J26:AB30
creator_box(ws, 26)

# ─────────────────────────────────────────────
# SHEET 2: 配偶者・子2人の場合
# ─────────────────────────────────────────────
ws = wb.create_sheet('配偶者・子2人の場合')
cw(ws)
rh(ws, {4:585,5:90,6:90,7:300,8:300,9:300,10:300,11:150,12:150,13:150,14:150,
        15:300,16:300,17:300,18:150,19:150,20:150,21:150,22:300,23:300,
        24:150,25:150,26:300,27:300,29:150,30:300,31:300,32:300,33:300,
        35:300,36:300,37:300,38:300,39:300,40:300,41:300,42:300})

mg(ws,'I4','M4','被相続人',F14,a=al(3))
yf(ws,4,14,4,20); mg(ws,'N4','T4')
mg(ws,'U4','AA4','法定相続情報',F14,a=al(2))

mg(ws,'A7','K7','最後の住所',a=al(1))
yf(ws,8,1,8,13); mg(ws,'A8','M8')
mg(ws,'A9','K9','最後の本籍　',a=al(1))
mg(ws,'P9','Q9','住所',a=al(1))
yf(ws,9,18,9,31); mg(ws,'R9','AE9')
yf(ws,10,1,10,13); mg(ws,'A10','M10')
mg(ws,'P10','Q10','出生  ',a=al(1))
yf(ws,10,18,10,27); mg(ws,'R10','AA10')

mg(ws,'A11','B12','出生  ',a=al(1))
yf(ws,11,3,12,11); mg(ws,'C11','K12')
yf(ws,11,16,12,18); mg(ws,'P11','R12','（　　）',a=al(1))

mg(ws,'A13','B14','死亡  ',a=al(1))
yf(ws,13,3,14,11); mg(ws,'C13','K14')
yf(ws,13,16,14,24); mg(ws,'P13','X14')
mg(ws,'Y13','AB14','(申出人)',a=al(2))

mg(ws,'A15','E15','（被相続人）',a=al(2))
yf(ws,16,1,16,9); mg(ws,'A16','I16')

# 子1 right side rows 20-27
mg(ws,'P20','Q21','住所',a=al(1))
yf(ws,20,18,21,31); mg(ws,'R20','AE21')
mg(ws,'A22','B22','住所　',a=al(1))
yf(ws,22,3,22,13); mg(ws,'C22','M22')
mg(ws,'P22','Q22','出生  ',a=al(1))
yf(ws,22,18,22,27); mg(ws,'R22','AA22')
mg(ws,'A23','B23','出生  ',a=al(0))
yf(ws,23,3,23,11); mg(ws,'C23','K23')
yf(ws,23,16,23,18); mg(ws,'P23','R23','（　　）',a=al(1))

yf(ws,24,1,25,5); mg(ws,'A24','E25','（　　）',a=al(1))
yf(ws,24,16,25,24); mg(ws,'P24','X25')
yf(ws,26,1,26,9); mg(ws,'A26','I26')
mg(ws,'P27','T27','以下余白',a=al(2))

# Creator rows 29-33
cv(ws,30,11,'作成日：',a=al(0))
yf(ws,30,14,30,24); mg(ws,'N30','X30')
cv(ws,31,11,'作成者：',a=al(0))
mg(ws,'N31','O31','住所',a=al(1))
yf(ws,31,16,31,28); mg(ws,'P31','AB31')
cv(ws,32,14,'氏名',a=al(2))
yf(ws,32,16,32,22); mg(ws,'P32','V32')

# Tree borders sheet 2
bd(ws,14,13,ri=1); bd(ws,14,14,t=1,l=1); bd(ws,14,15,t=1)
bd(ws,15,13,ri=1); bd(ws,15,14,l=1)
bd(ws,16,13,ri=1); bd(ws,16,14,l=1); bd(ws,17,3,l=6)
bd(ws,17,13,ri=1); bd(ws,17,14,l=1); bd(ws,18,3,l=6)
bd(ws,18,13,ri=1); bd(ws,18,14,l=1)
bd(ws,19,3,t=1,l=6)
for c in range(4,13): bd(ws,19,c,t=1)
bd(ws,19,13,t=1,ri=1); bd(ws,19,14,l=1)
bd(ws,20,3,l=6); bd(ws,20,13,ri=1); bd(ws,20,14,l=1)
bd(ws,21,3,l=6); bd(ws,21,13,ri=1); bd(ws,21,14,l=1)
bd(ws,22,13,ri=1); bd(ws,22,14,l=1)
bd(ws,23,13,ri=1); bd(ws,23,14,l=1)
bd(ws,24,13,ri=1); bd(ws,24,14,t=1,l=1); bd(ws,24,15,t=1)

creator_box(ws, 29)

# ─────────────────────────────────────────────
# SHEET 3: 配偶者・子3人の場合
# ─────────────────────────────────────────────
ws = wb.create_sheet('配偶者・子3人の場合')
cw(ws)
rh(ws, {4:585,5:90,6:300,7:300,8:300,9:300,10:300,11:150,12:150,13:150,14:150,
        15:300,16:300,17:150,18:150,19:150,20:150,21:150,22:150,23:300,
        24:150,25:150,26:300,27:300,28:150,29:150,30:300,31:300,33:150,
        34:300,35:300,36:300,37:300,39:300,40:300,41:300,42:300,43:300,
        44:300,45:300,46:300})

mg(ws,'I4','M4','被相続人',F14,a=al(3))
yf(ws,4,14,4,20); mg(ws,'N4','T4')
mg(ws,'U4','AA4','法定相続情報',F14,a=al(2))

mg(ws,'A6','E6','最後の住所',a=al(1))
yf(ws,7,1,7,12); mg(ws,'A7','L7')
mg(ws,'A8','E8','最後の本籍',a=al(1))
yf(ws,9,1,9,12); mg(ws,'A9','L9')
mg(ws,'P9','Q9','住所',a=al(1))
yf(ws,9,18,9,31); mg(ws,'R9','AE9')
mg(ws,'A10','B10','出生  ',a=al(1))
yf(ws,10,3,10,11); mg(ws,'C10','K10')
mg(ws,'P10','Q10','出生  ',a=al(1))
yf(ws,10,18,10,25); mg(ws,'R10','Y10')
mg(ws,'A11','B12','死亡  ',a=al(1))
yf(ws,11,3,12,11); mg(ws,'C11','K12')
yf(ws,11,16,12,18); mg(ws,'P11','R12','（　　）',a=al(1))
yf(ws,11,25,12,28); mg(ws,'Y11','AB12')

yf(ws,13,1,14,5); mg(ws,'A13','E14','（被相続人）',a=al(1))
yf(ws,13,16,14,24); mg(ws,'P13','X14')
mg(ws,'Y13','AB14','(申出人)',a=al(1))
yf(ws,15,1,15,9); mg(ws,'A15','I15')

# 子1 right rows 16-22
mg(ws,'P16','Q16','住所',a=al(1))
yf(ws,16,18,16,31); mg(ws,'R16','AE16')
mg(ws,'P17','Q18','出生',a=al(1))
yf(ws,17,18,18,25); mg(ws,'R17','Y18')
yf(ws,19,16,20,18); mg(ws,'P19','R20','（　　）',a=al(1))
yf(ws,21,1,22,2); mg(ws,'A21','B22','住所　',a=al(1))
yf(ws,21,3,22,12); mg(ws,'C21','L22')
yf(ws,21,16,22,24); mg(ws,'P21','X22')
mg(ws,'A23','B23','出生  ',a=al(0))
yf(ws,23,3,23,11); mg(ws,'C23','K23')

# 子2 left rows 24-28
yf(ws,24,1,25,5); mg(ws,'A24','E25','（　　）',a=al(1))
mg(ws,'P24','Q25','住所',a=al(1))
yf(ws,24,18,25,31); mg(ws,'R24','AE25')
yf(ws,26,1,26,9); mg(ws,'A26','I26')
mg(ws,'P26','Q26','出生',a=al(1))
yf(ws,26,18,26,25); mg(ws,'R26','Y26')
yf(ws,27,16,27,18); mg(ws,'P27','R27','（　　）',a=al(1))
yf(ws,28,16,29,24); mg(ws,'P28','X29')
mg(ws,'P31','T31','以下余白',a=al(2))

# Creator rows 33-37
cv(ws,34,11,'作成日：',a=al(0))
yf(ws,34,14,34,24); mg(ws,'N34','X34')
cv(ws,35,11,'作成者：',a=al(0))
mg(ws,'N35','O35','住所',a=al(1))
yf(ws,35,16,35,28); mg(ws,'P35','AB35')
cv(ws,36,14,'氏名',a=al(2))
yf(ws,36,16,36,22); mg(ws,'P36','V36')

# Tree borders sheet 3
bd(ws,13,13,b=1); bd(ws,13,14,b=1); bd(ws,13,15,b=1)
bd(ws,14,13,t=1,l=1); bd(ws,14,14,t=1); bd(ws,14,15,t=1)
bd(ws,15,13,l=1)
bd(ws,16,3,l=6); bd(ws,16,13,l=1)
bd(ws,17,3,l=6); bd(ws,17,13,l=1)
bd(ws,18,3,t=1,l=6)
for c in range(4,13): bd(ws,18,c,t=1)
bd(ws,18,13,l=1)
bd(ws,19,3,l=6); bd(ws,19,13,l=1)
bd(ws,20,3,l=6); bd(ws,20,13,l=1)
bd(ws,21,12,ri=1); bd(ws,21,13,l=1)
bd(ws,22,12,ri=1); bd(ws,22,13,t=1,l=1); bd(ws,22,14,t=1); bd(ws,22,15,t=1)
for r in range(23,28): bd(ws,r,13,l=1)
bd(ws,28,13,b=1,l=1); bd(ws,28,14,b=1); bd(ws,28,15,b=1)

creator_box(ws, 33)

# ─────────────────────────────────────────────
# SHEET 4: 配偶者・子4人の場合
# ─────────────────────────────────────────────
ws = wb.create_sheet('配偶者・子4人の場合')
cw(ws)
rh(ws, {4:585,5:90,6:300,7:300,8:300,9:300,10:300,11:150,12:150,13:150,14:150,
        15:300,16:300,17:150,18:150,19:150,20:150,21:150,22:150,23:300,
        24:150,25:150,26:300,27:300,28:150,29:150,30:300,31:300,32:300,33:300,
        34:150,35:150,36:300,37:300,38:300,39:150,40:300,41:300,42:300,43:300,
        45:300,46:300,47:300,48:300,49:300,50:300,51:300,52:300})

mg(ws,'I4','M4','被相続人',F14,a=al(3))
yf(ws,4,14,4,20); mg(ws,'N4','T4')
mg(ws,'U4','AA4','法定相続情報',F14,a=al(2))

mg(ws,'A6','E6','最後の住所',a=al(1))
yf(ws,7,1,7,12); mg(ws,'A7','L7')
mg(ws,'A8','E8','最後の本籍',a=al(1))
yf(ws,9,1,9,12); mg(ws,'A9','L9')
mg(ws,'P9','Q9','住所',a=al(1))
yf(ws,9,18,9,31); mg(ws,'R9','AE9')
mg(ws,'A10','B10','出生  ',a=al(1))
yf(ws,10,3,10,11); mg(ws,'C10','K10')
mg(ws,'P10','Q10','出生  ',a=al(1))
yf(ws,10,18,10,25); mg(ws,'R10','Y10')
mg(ws,'A11','B12','死亡  ',a=al(1))
yf(ws,11,3,12,11); mg(ws,'C11','K12')
yf(ws,11,16,12,18); mg(ws,'P11','R12','（　　）',a=al(1))
yf(ws,11,25,12,28); mg(ws,'Y11','AB12')

yf(ws,13,1,14,5); mg(ws,'A13','E14','（被相続人）',a=al(1))
yf(ws,13,16,14,24); mg(ws,'P13','X14')
mg(ws,'Y13','AB14','(申出人)',a=al(1))
yf(ws,15,1,15,9); mg(ws,'A15','I15')
mg(ws,'P15','Q15')
yf(ws,15,18,15,27); mg(ws,'R15','AA15')

# 子1 right rows 16-22
mg(ws,'P16','Q16','住所',a=al(1))
yf(ws,16,18,16,31); mg(ws,'R16','AE16')
mg(ws,'P17','Q18','出生',a=al(1))
yf(ws,17,18,18,25); mg(ws,'R17','Y18')
yf(ws,19,16,20,18); mg(ws,'P19','R20','（　　）',a=al(1))
yf(ws,21,1,22,2); mg(ws,'A21','B22','住所　',a=al(1))
yf(ws,21,3,22,12); mg(ws,'C21','L22')
yf(ws,21,16,22,24); mg(ws,'P21','X22')
mg(ws,'A23','B23','出生  ',a=al(0))
yf(ws,23,3,23,11); mg(ws,'C23','K23')

# 子2 left + 子3 right rows 24-30
yf(ws,24,1,25,5); mg(ws,'A24','E25','（　　）',a=al(1))
mg(ws,'P24','Q25','住所',a=al(1))
yf(ws,24,18,25,31); mg(ws,'R24','AE25')
yf(ws,26,1,26,9); mg(ws,'A26','I26')
mg(ws,'P26','Q26','出生',a=al(1))
yf(ws,26,18,26,25); mg(ws,'R26','Y26')
yf(ws,27,16,27,18); mg(ws,'P27','R27','（　　）',a=al(1))
yf(ws,28,16,29,24); mg(ws,'P28','X29')

# 子4 right rows 31-36
mg(ws,'P31','Q31','住所',a=al(1))
yf(ws,31,18,31,31); mg(ws,'R31','AE31')
mg(ws,'P32','Q32','出生',a=al(1))
yf(ws,32,18,32,25); mg(ws,'R32','Y32')
yf(ws,33,16,33,18); mg(ws,'P33','R33','（　　）',a=al(1))
yf(ws,34,16,35,24); mg(ws,'P34','X35')
mg(ws,'P37','T37','以下余白',a=al(2))

# Creator rows 39-43
cv(ws,40,11,'作成日：',a=al(0))
yf(ws,40,14,40,24); mg(ws,'N40','X40')
cv(ws,41,11,'作成者：',a=al(0))
mg(ws,'N41','O41','住所',a=al(1))
yf(ws,41,16,41,28); mg(ws,'P41','AB41')
cv(ws,42,14,'氏名',a=al(2))
yf(ws,42,16,42,22); mg(ws,'P42','V42')

# Tree borders sheet 4 (same as sheet 3 + extra)
bd(ws,13,13,b=1); bd(ws,13,14,b=1); bd(ws,13,15,b=1)
bd(ws,14,13,t=1,l=1); bd(ws,14,14,t=1); bd(ws,14,15,t=1)
bd(ws,15,13,l=1)
bd(ws,16,3,l=6); bd(ws,16,13,l=1)
bd(ws,17,3,l=6); bd(ws,17,13,l=1)
bd(ws,18,3,t=1,l=6)
for c in range(4,13): bd(ws,18,c,t=1)
bd(ws,18,13,l=1)
bd(ws,19,3,l=6); bd(ws,19,13,l=1)
bd(ws,20,3,l=6); bd(ws,20,13,l=1)
bd(ws,21,12,ri=1); bd(ws,21,13,l=1)
bd(ws,22,12,ri=1); bd(ws,22,13,t=1,l=1); bd(ws,22,14,t=1); bd(ws,22,15,t=1)
for r in range(23,28): bd(ws,r,13,l=1)
bd(ws,28,13,b=1,l=1); bd(ws,28,14,b=1); bd(ws,28,15,b=1)
bd(ws,29,13,t=1,l=1); bd(ws,29,14,t=1); bd(ws,29,15,t=1)
for r in range(30,34): bd(ws,r,13,l=1)
bd(ws,34,13,b=1,l=1); bd(ws,34,14,b=1); bd(ws,34,15,b=1)
for c in range(16,25): bd(ws,38,c,b=1)

creator_box(ws, 39)

# Save
out = "/Users/hamurakouichi/クロードテスト用/アプリ開発/法定相続情報一覧図　 Ver.2/配偶者・子（１人～４人まで対応）である場合.xlsx"
wb.save(out)
print(f"保存完了: {out}")
